#!/usr/bin/env python3
"""Local spreadsheet inspection, recalculation, rendering, and verification runtime."""

from __future__ import annotations

import argparse
import csv
import importlib.util
import json
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

CAPABILITIES = ["create", "read", "edit", "analyze", "formulas", "charts", "styles", "convert", "inspect", "recalculate", "render", "verify"]
MODULES = {"openpyxl": "openpyxl", "pandas": "pandas", "numpy": "numpy", "PIL": "Pillow"}
WORKBOOK_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
TABULAR_SUFFIXES = {".csv", ".tsv"}
FORMULA_ERRORS = ("#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A")


def emit(payload: dict[str, object], code: int = 0) -> int:
    print(json.dumps(payload, ensure_ascii=False))
    return code


def office_binary() -> str | None:
    executable_path = Path(sys.executable).resolve()
    runtime_root = executable_path.parents[2] if len(executable_path.parents) > 2 else executable_path.parent
    candidates = [
        os.environ.get("ONMYAGENT_LIBREOFFICE_PATH"),
        str(runtime_root / "bin" / "soffice"),
        str(runtime_root / "libreoffice" / "LibreOffice.app" / "Contents" / "MacOS" / "soffice"),
        str(runtime_root / "libreoffice" / "LibreOffice" / "program" / "soffice.exe"),
    ]
    for name in ("soffice", "libreoffice"):
        if binary := shutil.which(name):
            candidates.append(binary)
    candidates.append("/Applications/LibreOffice.app/Contents/MacOS/soffice")
    for candidate in candidates:
        if not candidate:
            continue
        resolved = str(Path(candidate).expanduser().resolve())
        if "/.cache/codex-runtimes/" in resolved or "/.codex/" in resolved:
            continue
        if Path(resolved).is_file():
            return resolved
    return None


def dependencies() -> dict[str, object]:
    modules = {package: importlib.util.find_spec(module) is not None for module, package in MODULES.items()}
    office = office_binary()
    return {"python_modules": modules, "office_renderer": {"available": office is not None, "path": office}, "ready_for_workbooks": all(modules.values()), "ready_for_recalculation": office is not None}


def require_sheet(source: Path) -> Path:
    source = source.resolve()
    if source.suffix.lower() not in WORKBOOK_SUFFIXES | TABULAR_SUFFIXES | {".xls"} or not source.is_file():
        raise ValueError(f"Input must be an existing XLSX, XLS, CSV, or TSV file: {source}")
    return source


def inspect_tabular(source: Path) -> dict[str, object]:
    delimiter = "\t" if source.suffix.lower() == ".tsv" else ","
    with source.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=delimiter))
    return {"status": "success", "source": str(source), "format": source.suffix.lower()[1:], "row_count": len(rows), "column_count": max((len(row) for row in rows), default=0), "sheets": []}


def inspect_workbook(source: Path, data_only: bool = False) -> dict[str, object]:
    if source.suffix.lower() == ".xls":
        raise RuntimeError("Legacy .xls inspection requires conversion with LibreOffice first")
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("openpyxl is unavailable; run the bundled runtime preparation step") from error
    workbook = load_workbook(source, read_only=True, data_only=data_only)
    sheets: list[dict[str, object]] = []
    formula_count = 0
    error_cells: list[dict[str, str]] = []
    try:
        for worksheet in workbook.worksheets:
            nonempty = 0
            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is not None:
                        nonempty += 1
                    if isinstance(value, str) and value.startswith("="):
                        formula_count += 1
                    if isinstance(value, str) and any(error in value for error in FORMULA_ERRORS):
                        error_cells.append({"sheet": worksheet.title, "cell": cell.coordinate, "value": value})
            sheets.append({"name": worksheet.title, "rows": worksheet.max_row, "columns": worksheet.max_column, "nonempty_cells": nonempty})
    finally:
        workbook.close()
    return {"status": "success", "source": str(source), "format": source.suffix.lower()[1:], "sheet_count": len(sheets), "sheets": sheets, "formula_count": formula_count, "formula_errors": error_cells[:100], "formula_error_count": len(error_cells)}


def inspect_sheet(source: Path, data_only: bool = False) -> dict[str, object]:
    source = require_sheet(source)
    return inspect_tabular(source) if source.suffix.lower() in TABULAR_SUFFIXES else inspect_workbook(source, data_only)


def run_office(source: Path, output_dir: Path, convert_to: str) -> Path:
    office = office_binary()
    if office is None:
        raise RuntimeError("LibreOffice is required for conversion, rendering, and recalculation but is unavailable")
    output_dir.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="onmyagent-sheet-") as temporary:
        profile = Path(temporary) / "profile"
        profile.mkdir()
        result = subprocess.run([office, "--headless", f"-env:UserInstallation={profile.as_uri()}", "--convert-to", convert_to, "--outdir", str(output_dir), str(source)], capture_output=True, text=True, timeout=120)
    if result.returncode != 0:
        detail = result.stderr.strip() or result.stdout.strip() or "unknown error"
        raise RuntimeError(f"LibreOffice failed: {detail}")
    extension = convert_to.split(":", 1)[0]
    output = output_dir / f"{source.stem}.{extension}"
    if not output.is_file():
        raise RuntimeError(f"LibreOffice completed without producing {output.name}")
    return output


def recalculate(source: Path, output_dir: Path) -> dict[str, object]:
    source = require_sheet(source)
    if source.suffix.lower() in TABULAR_SUFFIXES:
        raise ValueError("CSV and TSV files do not contain recalculable formulas")
    output = run_office(source, output_dir.resolve(), "xlsx")
    verification = inspect_workbook(output, data_only=True)
    status = "success" if verification["formula_error_count"] == 0 else "errors_found"
    return {"status": status, "source": str(source), "output": str(output), "verification": verification}


def render(source: Path, output_dir: Path) -> dict[str, object]:
    source = require_sheet(source)
    pdf = run_office(source, output_dir.resolve(), "pdf")
    return {"status": "success", "source": str(source), "pdf": str(pdf), "output_dir": str(output_dir.resolve())}


def verify(source: Path) -> dict[str, object]:
    source = require_sheet(source)
    if source.suffix.lower() in TABULAR_SUFFIXES:
        inspection = inspect_tabular(source)
        return {"status": "success", "inspection": inspection, "issues": []}
    formula_inspection = inspect_workbook(source, data_only=False)
    value_inspection = inspect_workbook(source, data_only=True)
    issues = []
    if value_inspection["formula_error_count"]:
        issues.append(f"{value_inspection['formula_error_count']} formula error cells found")
    if formula_inspection["formula_count"]:
        from openpyxl import load_workbook
        formulas = load_workbook(source, read_only=True, data_only=False)
        values = load_workbook(source, read_only=True, data_only=True)
        missing_cached_values: list[str] = []
        try:
            for formula_sheet, value_sheet in zip(formulas.worksheets, values.worksheets):
                for formula_row, value_row in zip(formula_sheet.iter_rows(), value_sheet.iter_rows()):
                    for formula_cell, value_cell in zip(formula_row, value_row):
                        if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") and value_cell.value is None:
                            missing_cached_values.append(f"{formula_sheet.title}!{formula_cell.coordinate}")
        finally:
            formulas.close()
            values.close()
        if missing_cached_values:
            issues.append(f"{len(missing_cached_values)} formula cells lack cached calculated values; run recalculate")
    return {
        "status": "success" if not issues else "issues_found",
        "inspection": formula_inspection,
        "calculated_values": value_inspection,
        "issues": issues,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("command", nargs="?", choices=("capabilities", "doctor", "inspect", "recalculate", "render", "verify"))
    parser.add_argument("input", nargs="?", type=Path)
    parser.add_argument("--output-dir", type=Path)
    parser.add_argument("--data-only", action="store_true")
    parser.add_argument("--capabilities", action="store_true")
    args = parser.parse_args()
    command = "capabilities" if args.capabilities else args.command
    try:
        if command == "capabilities":
            return emit({"status": "ready", "runtime": "spreadsheets", "capabilities": CAPABILITIES, "commands": ["doctor", "inspect", "recalculate", "render", "verify"]})
        if command == "doctor":
            report = dependencies()
            return emit({"status": "ready" if report["ready_for_workbooks"] else "degraded", "runtime": "spreadsheets", "dependencies": report, "capabilities": CAPABILITIES})
        if args.input is None:
            raise ValueError(f"{command or 'command'} requires an input path")
        if command == "inspect":
            return emit(inspect_sheet(args.input, args.data_only))
        if command in {"recalculate", "render"}:
            if args.output_dir is None:
                raise ValueError(f"{command} requires --output-dir")
            return emit(recalculate(args.input, args.output_dir) if command == "recalculate" else render(args.input, args.output_dir))
        if command == "verify":
            return emit(verify(args.input))
        raise ValueError("A command is required")
    except (OSError, RuntimeError, ValueError) as error:
        return emit({"status": "error", "runtime": "spreadsheets", "error": str(error)}, 1)


if __name__ == "__main__":
    raise SystemExit(main())
