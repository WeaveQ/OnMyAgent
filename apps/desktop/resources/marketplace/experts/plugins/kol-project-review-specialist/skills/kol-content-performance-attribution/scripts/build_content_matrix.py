#!/usr/bin/env python3
"""Build weekly and evidence matrices for creator content performance."""

from __future__ import annotations

import argparse
import json
import math
import sys
from collections import defaultdict
from pathlib import Path


ALIASES = {
    "record_id": ("record_id", "记录ID", "笔记ID", "发布链接"),
    "week": ("week", "周次", "周期"),
    "content_type": ("content_type", "内容类型", "笔记类型"),
    "content_angle": ("content_angle", "内容切角", "切角标签"),
    "audience": ("audience", "目标人群", "人群"),
    "search_term": ("search_term", "搜索词", "关键词"),
    "action": ("action", "优化动作", "投放动作", "调整动作"),
    "spend": ("spend", "投入", "消耗", "广告金额"),
    "reads": ("reads", "阅读量", "推广阅读量"),
    "interactions": ("interactions", "互动量", "赞评藏"),
}


def number(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        value = value.replace(",", "").replace("¥", "").replace("￥", "").strip()
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def load_json(value: str | None) -> dict[str, object]:
    if not value:
        return {}
    if value.lstrip().startswith("{"):
        content = value
    else:
        path = Path(value)
        content = path.read_text(encoding="utf-8") if path.exists() else value
    parsed = json.loads(content)
    if not isinstance(parsed, dict):
        raise ValueError("列映射必须是 JSON 对象")
    return parsed


def text_value(value: object) -> str:
    return "" if value is None else str(value).strip()


def excel_safe(value: object) -> object:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def append_safe(sheet: object, values: list[object]) -> None:
    sheet.append([excel_safe(value) for value in values])


def map_headers(headers: list[str], custom: dict[str, object]) -> dict[str, list[int]]:
    mapping: dict[str, list[int]] = {}
    for canonical, defaults in ALIASES.items():
        configured = custom.get(canonical, [])
        candidates = list(defaults) + ([str(configured)] if isinstance(configured, str) else [str(item) for item in configured])
        indexes = list(dict.fromkeys(index for candidate in candidates for index, header in enumerate(headers) if header == candidate))
        if indexes:
            mapping[canonical] = indexes
    return mapping


def aggregate(records: list[dict[str, object]], keys: tuple[str, ...]) -> list[dict[str, object]]:
    groups: defaultdict[tuple[str, ...], list[dict[str, object]]] = defaultdict(list)
    for record in records:
        key = tuple(text_value(record.get(field)) or "未提供" for field in keys)
        groups[key].append(record)
    output: list[dict[str, object]] = []
    for key, items in groups.items():
        spend = sum(value for item in items if (value := number(item.get("spend"))) is not None)
        reads = sum(value for item in items if (value := number(item.get("reads"))) is not None)
        interactions = sum(value for item in items if (value := number(item.get("interactions"))) is not None)
        row: dict[str, object] = {field: key[index] for index, field in enumerate(keys)}
        row.update({
            "源行": "、".join(str(item["源行"]) for item in items),
            "样本数": len(items),
            "投入": spend,
            "阅读量": reads,
            "互动量": interactions,
            "结论类型": "相关性观察",
            "动作记录": "；".join(
                dict.fromkeys(text_value(item.get("action")) for item in items if text_value(item.get("action")))
            ),
        })
        output.append(row)
    return output


def _append_matrix(
    sheet, keys: tuple[str, ...], rows: list[dict[str, object]], include_actions: bool = False
) -> None:
    headers = [*keys, "源行", "样本数", "投入", "阅读量", "互动量"]
    if include_actions:
        headers.append("动作记录")
    headers.extend(["结论类型", "风险提示"])
    append_safe(sheet, headers)
    for row in rows:
        risks = []
        if row["样本数"] < 3:
            risks.append("样本不足")
        values = [
            *[row[key] for key in keys],
            row["源行"],
            row["样本数"],
            row["投入"],
            row["阅读量"],
            row["互动量"],
        ]
        if include_actions:
            values.append(row["动作记录"])
        values.extend([row["结论类型"], "；".join(risks)])
        append_safe(sheet, values)


def build(args: argparse.Namespace) -> dict[str, int]:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as error:
        raise RuntimeError("缺少 openpyxl，无法读写 Excel；请使用桌面应用打包运行时") from error

    input_path = Path(args.input)
    if not input_path.exists():
        raise ValueError("投放明细文件不存在")
    source = load_workbook(input_path, read_only=True, data_only=True)
    if args.sheet and args.sheet not in source.sheetnames:
        source.close()
        raise ValueError(f"输入工作表不存在：{args.sheet}")
    sheet = source[args.sheet] if args.sheet else source.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        first_row = next(iterator)
    except StopIteration as error:
        source.close()
        raise ValueError("输入工作表为空") from error
    headers = [text_value(value) for value in first_row]
    mapping = map_headers(headers, load_json(args.column_map))
    required = {"week", "content_type", "audience", "search_term", "spend", "reads", "interactions"}
    missing_columns = sorted(required.difference(mapping))
    if missing_columns:
        source.close()
        raise ValueError("缺少必要列：" + "、".join(missing_columns))
    records: list[dict[str, object]] = []
    duplicate_counts: defaultdict[str, int] = defaultdict(int)
    for source_row, values in enumerate(iterator, start=2):
        record: dict[str, object] = {}
        field_conflicts: list[str] = []
        for key, indexes in mapping.items():
            candidates = [values[index] for index in indexes if index < len(values) and values[index] not in (None, "")]
            unique = list(dict.fromkeys(str(value).strip() for value in candidates))
            record[key] = candidates[0] if candidates else None
            if len(unique) > 1:
                field_conflicts.append(key)
        record["源行"] = source_row
        record["字段冲突"] = field_conflicts
        record["缺失数值字段"] = [
            field for field in ("spend", "reads", "interactions") if number(record.get(field)) is None
        ]
        record_id = text_value(record.get("record_id")) or f"源行{source_row}"
        record["record_id"] = record_id
        duplicate_counts[record_id] += 1
        records.append(record)
    source.close()

    valid_records = [
        record
        for record in records
        if duplicate_counts[str(record["record_id"])] == 1
        and not record["字段冲突"]
        and not record["缺失数值字段"]
    ]
    weekly = aggregate(valid_records, ("week",))
    content_audience = aggregate(valid_records, ("content_type", "audience"))
    search_content = aggregate(valid_records, ("search_term", "content_type"))
    content_angle = aggregate(valid_records, ("content_angle", "audience"))

    workbook = Workbook()
    weekly_sheet = workbook.active
    weekly_sheet.title = "周环比"
    _append_matrix(weekly_sheet, ("week",), weekly, include_actions=True)
    _append_matrix(workbook.create_sheet("内容人群矩阵"), ("content_type", "audience"), content_audience)
    _append_matrix(workbook.create_sheet("搜索内容矩阵"), ("search_term", "content_type"), search_content)
    angle_sheet = workbook.create_sheet("内容切角")
    _append_matrix(angle_sheet, ("content_angle", "audience"), content_angle)

    risks = workbook.create_sheet("风险提示")
    append_safe(risks, ["类型", "对象", "说明", "处理"])
    for record_id, count in duplicate_counts.items():
        if count > 1:
            append_safe(risks, ["重复记录", record_id, f"出现 {count} 次", "待人工确认，未进入矩阵"])
    for record in records:
        if record["字段冲突"]:
            append_safe(risks, ["字段冲突", record["record_id"], "、".join(record["字段冲突"]), "待人工确认，未进入矩阵"])
        if record["缺失数值字段"]:
            append_safe(
                risks,
                [
                    "数值缺失或无效",
                    record["record_id"],
                    "、".join(record["缺失数值字段"]),
                    "标记数据不足，未进入矩阵",
                ],
            )
    for matrix_name, rows in (("内容×人群", content_audience), ("搜索词×内容", search_content), ("内容切角", content_angle)):
        for row in rows:
            if row["样本数"] < 3:
                label = " / ".join(str(row[key]) for key in row if key in {"content_type", "audience", "search_term", "content_angle"})
                append_safe(risks, ["样本不足", f"{matrix_name}: {label}", f"样本数 {row['样本数']}", "仅作相关性观察"])
    append_safe(risks, ["结论边界", "全部矩阵", "相关性不等于因果", "需要实验或额外证据才能提出因果结论"])

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {"records": len(records), "content_audience_cells": len(content_audience), "search_content_cells": len(search_content)}


def self_test() -> dict[str, object]:
    sample = [
        {"源行": 2, "content_type": "测评", "audience": "新客", "search_term": "成分", "spend": 100, "reads": 1000, "interactions": 100},
        {"源行": 3, "content_type": "教程", "audience": "熟客", "search_term": "用法", "spend": 120, "reads": 800, "interactions": 80},
    ]
    return {
        "content_audience_cells": len(aggregate(sample, ("content_type", "audience"))),
        "search_content_cells": len(aggregate(sample, ("search_term", "content_type"))),
        "conclusion_kind": aggregate(sample, ("content_type", "audience"))[0]["结论类型"],
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="生成内容效能周环比和矩阵")
    parser.add_argument("--input")
    parser.add_argument("--output")
    parser.add_argument("--sheet")
    parser.add_argument("--column-map")
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args()
    if args.self_test:
        print(json.dumps(self_test(), ensure_ascii=False))
        return 0
    if not args.input or not args.output:
        print("必须提供 --input 和 --output", file=sys.stderr)
        return 2
    try:
        result = build(args)
    except (OSError, ValueError, RuntimeError, StopIteration) as error:
        print(f"生成失败：{error}", file=sys.stderr)
        return 1
    print(json.dumps(result, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
