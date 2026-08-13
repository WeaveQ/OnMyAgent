#!/usr/bin/env python3
"""Create a traceable KOL margin and performance workbook."""

from __future__ import annotations

import argparse
import html
import json
import math
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path


ALIASES = {
    "record_id": ("record_id", "记录ID", "笔记ID", "笔记链接", "发布链接"),
    "creator": ("creator", "达人", "达人名称", "博主", "达人昵称", "博主昵称"),
    "rebate_band": ("rebate_band", "返点档", "返点档位"),
    "margin_rate": ("margin_rate", "毛利率"),
    "creator_tier": ("creator_tier", "达人量级", "达人等级", "粉丝量级"),
    "note_type": ("note_type", "笔记类型", "内容类型"),
    "k_amount": ("k_amount", "K金额", "k金额", "K 金额"),
    "ad_amount": ("ad_amount", "广告金额", "投流金额"),
    "natural_reads": ("natural_reads", "自然阅读量", "自然阅读"),
    "natural_impressions": ("natural_impressions", "自然曝光量", "自然曝光"),
    "likes": ("likes", "点赞量", "点赞"),
    "comments": ("comments", "评论量", "评论"),
    "favorites": ("favorites", "收藏量", "收藏"),
    "search_uv": ("search_uv", "搜索进店UV", "搜索进店 UV"),
}
PLAN_KEY_ALIASES = ("笔记链接", "发布链接", "记录ID", "笔记ID")
EFFECT_KEY_ALIASES = ("发布链接", "笔记链接", "记录ID", "笔记ID")
PLAN_FIELD_ALIASES = {
    "project": ("项目", "项目名称"),
    "record_id": PLAN_KEY_ALIASES,
    "creator": ("达人昵称", "达人", "达人名称", "博主"),
    "rebate_band": ("返点档", "返点档位"),
    "margin_rate": ("毛利率",),
    "creator_tier": ("粉丝量级", "达人量级", "达人等级"),
    "note_type": ("笔记类型", "内容类型"),
    "k_amount": ("K金额", "K 金额", "k金额"),
}
EFFECT_FIELD_ALIASES = {
    "effect_record_id": ("效果记录ID", "记录ID", "笔记ID"),
    "effect_creator": ("博主昵称", "达人昵称", "达人", "博主"),
    "ad_amount": ("广告金额", "投流金额"),
    "natural_reads": ("自然阅读量", "自然阅读"),
    "natural_impressions": ("自然曝光量", "自然曝光"),
    "likes": ("点赞量", "点赞"),
    "comments": ("评论量", "评论"),
    "favorites": ("收藏量", "收藏"),
    "search_uv": ("搜索进店UV", "搜索进店 UV"),
    "promoted_impressions": ("推广曝光量", "广告曝光量"),
    "promoted_reads": ("推广阅读量", "广告阅读量"),
    "content_angle": ("内容切角",),
    "target_audience": ("目标人群",),
    "search_term": ("搜索词",),
    "week": ("周次",),
}
FORMULAS = (
    ("总金额", "K 金额 + 广告金额"),
    ("自然 CTR", "自然阅读量 / 自然曝光量"),
    ("纯 K CPC", "K 金额 / 自然阅读量"),
    ("互动量", "点赞量 + 评论量 + 收藏量"),
    ("广告占比", "广告金额 / 总金额"),
    ("搜索进店成本", "总金额 / 搜索进店 UV"),
)


def number(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace("¥", "").replace("￥", "").strip()
        if cleaned.endswith("%"):
            try:
                return float(cleaned[:-1]) / 100
            except ValueError:
                return None
        value = cleaned
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def safe_div(numerator: float | None, denominator: float | None) -> float | None:
    if numerator is None or denominator is None or denominator == 0:
        return None
    return numerator / denominator


def metrics(row: dict[str, object]) -> dict[str, float | None]:
    k_amount = number(row.get("k_amount"))
    ad_amount = number(row.get("ad_amount"))
    natural_reads = number(row.get("natural_reads"))
    natural_impressions = number(row.get("natural_impressions"))
    likes = number(row.get("likes"))
    comments = number(row.get("comments"))
    favorites = number(row.get("favorites"))
    search_uv = number(row.get("search_uv"))
    total_amount = k_amount + ad_amount if k_amount is not None and ad_amount is not None else None
    interaction_count = likes + comments + favorites if None not in (likes, comments, favorites) else None
    return {
        "总金额": total_amount,
        "自然 CTR": safe_div(natural_reads, natural_impressions),
        "纯 K CPC": safe_div(k_amount, natural_reads),
        "互动量": interaction_count,
        "广告占比": safe_div(ad_amount, total_amount),
        "搜索进店成本": safe_div(total_amount, search_uv),
    }


def row_label(values: dict[str, float | None], thresholds: dict[str, object] | None) -> str:
    if not thresholds or any(value is None for value in values.values()):
        return "数据不足"
    min_ctr = number(thresholds.get("min_natural_ctr"))
    max_cpc = number(thresholds.get("max_pure_k_cpc"))
    max_search = number(thresholds.get("max_search_visit_cost"))
    if None in (min_ctr, max_cpc, max_search):
        return "数据不足"
    min_interactions = number(thresholds.get("min_interactions"))
    interaction_good = min_interactions is None or values["互动量"] >= min_interactions
    good = values["自然 CTR"] >= min_ctr and values["纯 K CPC"] <= max_cpc and values["搜索进店成本"] <= max_search and interaction_good
    poor = values["自然 CTR"] < min_ctr and values["纯 K CPC"] > max_cpc
    if good:
        return "优先保留"
    if poor:
        return "谨慎"
    return "可加测"


def load_json(value: str | None) -> dict[str, object]:
    if not value:
        return {}
    if value.lstrip().startswith("{"):
        content = value
    else:
        path = Path(value)
        content = path.read_text(encoding="utf-8") if path.exists() else value
    parsed = json.loads(content)
    if not isinstance(parsed, dict):
        raise ValueError("配置必须是 JSON 对象")
    return parsed


def first_header(headers: list[str], candidates: tuple[str, ...]) -> str | None:
    return next((candidate for candidate in candidates if candidate in headers), None)


def first_value(row: dict[str, object], candidates: tuple[str, ...]) -> object:
    for candidate in candidates:
        value = row.get(candidate)
        if value not in (None, ""):
            return value
    return None


def read_rows(path: Path, sheet_name: str | None) -> tuple[str, list[str], list[dict[str, object]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("缺少 openpyxl，无法读写 Excel；请使用桌面应用打包运行时") from error

    if not path.exists():
        raise ValueError(f"项目数据文件不存在：{path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name and sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"输入工作表不存在：{sheet_name}")
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    values = sheet.iter_rows(values_only=True)
    try:
        headers = [text_value(value) for value in next(values)]
    except StopIteration as error:
        workbook.close()
        raise ValueError(f"输入工作表为空：{sheet.title}") from error
    rows = [
        {"__source_row": source_row, **dict(zip(headers, row_values))}
        for source_row, row_values in enumerate(values, start=2)
    ]
    title = sheet.title
    workbook.close()
    return title, headers, rows


def parse_threshold(value: object) -> float | None:
    if value in (None, ""):
        return None
    raw = str(value).strip().replace(">=", "").replace("<=", "").replace("≥", "").replace("≤", "")
    return number(raw)


def read_method_sheet(path: Path, sheet_name: str | None) -> tuple[list[list[object]], dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("缺少 openpyxl，无法读写 Excel；请使用桌面应用打包运行时") from error

    workbook = load_workbook(path, read_only=True, data_only=True)
    target = sheet_name
    if target is None:
        target = next((name for name in workbook.sheetnames if "口径" in name or "阈值" in name), None)
    if target is None:
        workbook.close()
        return [], {}
    if target not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"口径工作表不存在：{target}")
    rows = [list(row) for row in workbook[target].iter_rows(values_only=True)]
    workbook.close()
    thresholds: dict[str, object] = {}
    for row in rows[1:]:
        metric = text_value(row[1] if len(row) > 1 else None).replace(" ", "").upper()
        threshold = parse_threshold(row[3] if len(row) > 3 else None)
        if threshold is None:
            continue
        if metric == "自然CTR":
            thresholds["min_natural_ctr"] = threshold
        elif metric == "纯KCPC":
            thresholds["max_pure_k_cpc"] = threshold
        elif metric == "互动量":
            thresholds["min_interactions"] = threshold
        elif metric == "搜索进店成本":
            thresholds["max_search_visit_cost"] = threshold
    return rows, thresholds


def index_by_key(rows: list[dict[str, object]], header: str) -> tuple[defaultdict[str, list[dict[str, object]]], list[dict[str, object]]]:
    indexed: defaultdict[str, list[dict[str, object]]] = defaultdict(list)
    missing: list[dict[str, object]] = []
    for row in rows:
        key = text_value(row.get(header))
        if key:
            indexed[key].append(row)
        else:
            missing.append(row)
    return indexed, missing


def prepare_two_source_records(args: argparse.Namespace) -> tuple[list[dict[str, object]], list[list[object]], list[list[object]], dict[str, object], dict[str, object]]:
    plan_path = Path(args.plan_input)
    effect_path = Path(args.effect_input)
    plan_sheet, plan_headers, plan_rows = read_rows(plan_path, args.plan_sheet)
    effect_sheet, effect_headers, effect_rows = read_rows(effect_path, args.effect_sheet)
    plan_key_header = first_header(plan_headers, PLAN_KEY_ALIASES)
    effect_key_header = first_header(effect_headers, EFFECT_KEY_ALIASES)
    if not plan_key_header or not effect_key_header:
        raise ValueError("双表模式必须能发现计划单与效果表的笔记链接字段")

    plan_index, missing_plan_keys = index_by_key(plan_rows, plan_key_header)
    effect_index, missing_effect_keys = index_by_key(effect_rows, effect_key_header)
    manual: list[list[object]] = []
    records: list[dict[str, object]] = []

    def add_manual(source: str, row: dict[str, object], key: str, reason: str) -> None:
        manual.append([source, row.get("__source_row"), key, reason, "人工确认"])

    for row in missing_plan_keys:
        add_manual(plan_path.name, row, "", "计划单缺少笔记链接")
    for row in missing_effect_keys:
        add_manual(effect_path.name, row, "", "效果表缺少笔记链接")

    all_keys = sorted(set(plan_index) | set(effect_index))
    for key in all_keys:
        plans = plan_index.get(key, [])
        effects = effect_index.get(key, [])
        if len(plans) != 1 or len(effects) != 1:
            if not plans:
                for row in effects:
                    add_manual(effect_path.name, row, key, "效果表记录在计划单中未匹配")
            elif not effects:
                for row in plans:
                    add_manual(plan_path.name, row, key, "计划单记录在效果表中未匹配")
            else:
                reason = f"笔记链接不是一对一：计划单 {len(plans)} 条，效果表 {len(effects)} 条"
                for row in plans:
                    add_manual(plan_path.name, row, key, reason)
                for row in effects:
                    add_manual(effect_path.name, row, key, reason)
            continue

        plan = plans[0]
        effect = effects[0]
        record: dict[str, object] = {
            "计划单源行": plan["__source_row"],
            "效果表源行": effect["__source_row"],
            "笔记链接": key,
        }
        for field, aliases in PLAN_FIELD_ALIASES.items():
            record[field] = first_value(plan, aliases)
        for field, aliases in EFFECT_FIELD_ALIASES.items():
            record[field] = first_value(effect, aliases)
        record["record_id"] = record.get("effect_record_id") or key
        record.update(metrics(record))
        records.append(record)

    method_rows, sheet_thresholds = read_method_sheet(plan_path, args.method_sheet)
    thresholds = {**sheet_thresholds, **load_json(args.thresholds)}
    source_info = {
        "plan_path": plan_path,
        "effect_path": effect_path,
        "plan_sheet": plan_sheet,
        "effect_sheet": effect_sheet,
        "plan_key_header": plan_key_header,
        "effect_key_header": effect_key_header,
        "plan_headers": plan_headers,
        "effect_headers": effect_headers,
        "plan_count": len(plan_rows),
        "effect_count": len(effect_rows),
    }
    return records, manual, method_rows, thresholds, source_info


def map_headers(headers: list[str], custom: dict[str, object]) -> dict[str, list[int]]:
    mapping: dict[str, list[int]] = {}
    for canonical, defaults in ALIASES.items():
        configured = custom.get(canonical, [])
        candidates = list(defaults) + ([str(configured)] if isinstance(configured, str) else [str(item) for item in configured])
        indexes = list(dict.fromkeys(index for candidate in candidates for index, header in enumerate(headers) if header == candidate))
        if indexes:
            mapping[canonical] = indexes
    return mapping


def text_value(value: object) -> str:
    return "" if value is None else str(value).strip()


def excel_safe(value: object) -> object:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def append_safe(sheet: object, values: list[object]) -> None:
    sheet.append([excel_safe(value) for value in values])


def average(values: list[float | None]) -> float | None:
    usable = [value for value in values if value is not None]
    return sum(usable) / len(usable) if usable else None


def format_number(value: float | None, *, kind: str = "number") -> str:
    if value is None:
        return "数据不足"
    if kind == "percent":
        return f"{value:.1%}"
    if kind == "currency":
        return f"¥{value:,.2f}"
    return f"{value:,.0f}"


def grouped_summaries(records: list[dict[str, object]]) -> list[dict[str, object]]:
    summaries: list[dict[str, object]] = []
    valid_records = [record for record in records if record.get("行级标签") != "数据不足"]
    for field, label in (("rebate_band", "返点档"), ("margin_rate", "毛利率"), ("creator_tier", "粉丝量级"), ("note_type", "笔记类型")):
        groups: defaultdict[str, list[dict[str, object]]] = defaultdict(list)
        for record in valid_records:
            groups[text_value(record.get(field)) or "未提供"].append(record)
        for group_name, items in groups.items():
            summaries.append({
                "dimension": label,
                "name": group_name,
                "count": len(items),
                "ctr": average([number(item.get("自然 CTR")) for item in items]),
                "cpc": average([number(item.get("纯 K CPC")) for item in items]),
                "interaction": average([number(item.get("互动量")) for item in items]),
                "ad_share": average([number(item.get("广告占比")) for item in items]),
                "search_cost": average([number(item.get("搜索进店成本")) for item in items]),
            })
    return summaries


def build_html_report(
    output_path: Path,
    records: list[dict[str, object]],
    groups: list[dict[str, object]],
    manual_rows: list[list[object]],
    source_info: dict[str, object],
    thresholds: dict[str, object],
) -> None:
    project = next((text_value(record.get("project")) for record in records if text_value(record.get("project"))), "达人营销项目")
    label_counts: defaultdict[str, int] = defaultdict(int)
    for record in records:
        label_counts[text_value(record.get("行级标签")) or "数据不足"] += 1
    matched = len(records)
    plan_count = int(source_info.get("plan_count", matched))
    effect_count = int(source_info.get("effect_count", matched))
    issues = len(manual_rows)
    avg_ctr = average([number(record.get("自然 CTR")) for record in records])
    avg_cpc = average([number(record.get("纯 K CPC")) for record in records])
    avg_interaction = average([number(record.get("互动量")) for record in records])
    avg_ad_share = average([number(record.get("广告占比")) for record in records])
    avg_search = average([number(record.get("搜索进店成本")) for record in records])
    label_order = {"优先保留": 0, "可加测": 1, "谨慎": 2, "数据不足": 3}
    ranked = sorted(
        records,
        key=lambda record: (
            label_order.get(text_value(record.get("行级标签")), 9),
            number(record.get("纯 K CPC")) if number(record.get("纯 K CPC")) is not None else float("inf"),
        ),
    )

    def e(value: object) -> str:
        return html.escape(text_value(value), quote=True)

    def status_class(label: str) -> str:
        return {"优先保留": "good", "可加测": "watch", "谨慎": "risk", "数据不足": "missing"}.get(label, "missing")

    def metric_card(label: str, value: str, note: str) -> str:
        return f'<article class="metric"><span>{e(label)}</span><strong>{e(value)}</strong><small>{e(note)}</small></article>'

    def sample_label(group: dict[str, object]) -> str:
        return '<span class="note">小样本</span>' if int(group["count"]) < 3 else "相关性观察"

    metric_cards = "".join([
        metric_card("一对一匹配", f"{matched} 条", f"计划 {plan_count} / 效果 {effect_count}"),
        metric_card("平均自然 CTR", format_number(avg_ctr, kind="percent"), "仅统计可计算样本"),
        metric_card("平均纯 K CPC", format_number(avg_cpc, kind="currency"), "K 金额 / 自然阅读量"),
        metric_card("平均搜索进店成本", format_number(avg_search, kind="currency"), "总金额 / 搜索进店 UV"),
    ])
    label_chips = "".join(
        f'<span class="chip {status_class(label)}">{e(label)} {label_counts.get(label, 0)}</span>'
        for label in ("优先保留", "可加测", "谨慎", "数据不足")
    )
    ranking_rows = "".join(
        "<tr>"
        f'<td><strong>{e(record.get("creator"))}</strong><small>{e(record.get("record_id"))}</small></td>'
        f'<td><span class="chip {status_class(text_value(record.get("行级标签")))}">{e(record.get("行级标签"))}</span></td>'
        f'<td>{e(format_number(number(record.get("自然 CTR")), kind="percent"))}</td>'
        f'<td>{e(format_number(number(record.get("纯 K CPC")), kind="currency"))}</td>'
        f'<td>{e(format_number(number(record.get("互动量"))))}</td>'
        f'<td>{e(format_number(number(record.get("广告占比")), kind="percent"))}</td>'
        f'<td>{e(format_number(number(record.get("搜索进店成本")), kind="currency"))}</td>'
        "</tr>"
        for record in ranked
    )
    group_rows = "".join(
        "<tr>"
        f'<td>{e(group["dimension"])}</td><td><strong>{e(group["name"])}</strong></td>'
        f'<td>{e(group["count"])}</td>'
        f'<td>{e(format_number(number(group["ctr"]), kind="percent"))}</td>'
        f'<td>{e(format_number(number(group["cpc"]), kind="currency"))}</td>'
        f'<td>{e(format_number(number(group["search_cost"]), kind="currency"))}</td>'
        f'<td>{sample_label(group)}</td>'
        "</tr>"
        for group in groups
    )
    manual_items = "".join(
        f'<li><strong>{e(row[2] or "缺少链接")}</strong><span>{e(row[3])}</span><small>{e(row[0])} · 源行 {e(row[1])}</small></li>'
        for row in manual_rows
    ) or "<li><strong>无匹配异常</strong><span>本次没有重复、一对多或未匹配记录。</span></li>"
    data_gap_items = "".join(
        f'<li><strong>{e(record.get("record_id"))}</strong><span>{e("、".join(name for name, _ in FORMULAS if record.get(name) is None))} 数据不足</span></li>'
        for record in records
        if any(record.get(name) is None for name, _ in FORMULAS)
    ) or "<li><strong>关键指标完整</strong><span>纳入分析的样本均可计算核心指标。</span></li>"
    threshold_text = " / ".join(
        filter(None, [
            f'自然 CTR ≥ {format_number(number(thresholds.get("min_natural_ctr")), kind="percent")}' if thresholds.get("min_natural_ctr") is not None else "",
            f'纯 K CPC ≤ {format_number(number(thresholds.get("max_pure_k_cpc")), kind="currency")}' if thresholds.get("max_pure_k_cpc") is not None else "",
            f'互动量 ≥ {format_number(number(thresholds.get("min_interactions")))}' if thresholds.get("min_interactions") is not None else "",
            f'搜索进店成本 ≤ {format_number(number(thresholds.get("max_search_visit_cost")), kind="currency")}' if thresholds.get("max_search_visit_cost") is not None else "",
        ])
    ) or "业务材料未提供完整阈值，涉及标签的结论须补充口径后复核。"
    conclusion = (
        f"本次按笔记链接精确得到 {matched} 条一对一样本；"
        f"其中 {label_counts.get('优先保留', 0)} 条优先保留、{label_counts.get('可加测', 0)} 条可加测，"
        f"另有 {issues} 项匹配或关键数据问题需要人工确认。"
    )
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    plan_name = Path(source_info["plan_path"]).name if source_info.get("plan_path") else "计划单"
    effect_name = Path(source_info["effect_path"]).name if source_info.get("effect_path") else "效果表"
    document = f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{e(project)} · 返点毛利与投放效果分析</title>
<style>
:root{{--ink:#17202a;--muted:#67717d;--paper:#f4f1e8;--panel:#fffdf8;--line:#ded8cc;--accent:#b25f37;--navy:#213f4f;--green:#27745d;--amber:#a66b1f;--red:#a3463c}}
*{{box-sizing:border-box}} body{{margin:0;background:var(--paper);color:var(--ink);font-family:"PingFang SC","Microsoft YaHei",system-ui,sans-serif;line-height:1.55}}
main{{max-width:1180px;margin:0 auto;padding:48px 32px 72px}} header{{display:grid;grid-template-columns:1.55fr .75fr;gap:48px;padding:34px 0 28px;border-top:8px solid var(--navy);border-bottom:1px solid var(--line)}}
.eyebrow{{font-size:12px;font-weight:700;letter-spacing:.16em;color:var(--accent)}} h1{{font-size:42px;line-height:1.15;margin:12px 0 16px;letter-spacing:-.03em}} .lead{{font-size:18px;max-width:780px;margin:0;color:#3d4852}} .meta{{align-self:end;color:var(--muted);font-size:13px}} .meta strong{{display:block;color:var(--ink);font-size:14px;margin-bottom:8px}}
.metrics{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin:26px 0}} .metric,.panel{{background:var(--panel);border:1px solid var(--line);border-radius:12px}} .metric{{padding:20px}} .metric span,.metric small{{display:block;color:var(--muted);font-size:12px}} .metric strong{{display:block;font:700 28px/1.2 ui-monospace,SFMono-Regular,Menlo,monospace;margin:8px 0;color:var(--navy)}}
.grid{{display:grid;grid-template-columns:1.45fr .75fr;gap:18px;margin-top:18px}} .panel{{padding:24px;overflow:hidden}} h2{{font-size:20px;margin:0 0 14px}} h3{{font-size:13px;color:var(--muted);margin:22px 0 8px;text-transform:uppercase;letter-spacing:.08em}} .chips{{display:flex;flex-wrap:wrap;gap:8px}}
.chip{{display:inline-flex;align-items:center;border-radius:99px;padding:4px 9px;font-size:12px;font-weight:700;background:#e8e9e7;color:#59616a}} .good{{background:#dcebe4;color:var(--green)}} .watch{{background:#f4e7cf;color:var(--amber)}} .risk{{background:#f1dbd8;color:var(--red)}} .missing{{background:#e5e7e7;color:#59616a}}
table{{width:100%;border-collapse:collapse;font-size:13px}} th{{text-align:left;padding:11px 10px;border-bottom:2px solid var(--navy);color:var(--muted);font-size:11px;letter-spacing:.04em}} td{{padding:12px 10px;border-bottom:1px solid var(--line);vertical-align:top}} td small{{display:block;color:var(--muted);margin-top:3px}} .scroll{{overflow-x:auto}}
.note{{display:inline-block;background:#f4e7cf;color:var(--amber);padding:2px 6px;border-radius:4px;font-size:11px;font-weight:700}} ul{{list-style:none;padding:0;margin:0}} li{{padding:11px 0;border-bottom:1px solid var(--line)}} li strong,li span,li small{{display:block}} li span{{color:#3d4852;margin-top:3px}} li small{{color:var(--muted);margin-top:4px}} .method{{color:var(--muted);font-size:13px}} footer{{margin-top:22px;color:var(--muted);font-size:12px;border-top:1px solid var(--line);padding-top:18px}}
@media(max-width:820px){{main{{padding:24px 16px}}header,.grid{{grid-template-columns:1fr}}h1{{font-size:32px}}.metrics{{grid-template-columns:repeat(2,1fr)}}}}
@media print{{body{{background:#fff}}main{{max-width:none;padding:20px}}.panel,.metric{{break-inside:avoid}}}}
</style></head><body><main>
<header><div><div class="eyebrow">PROJECT REVIEW · 决策摘要</div><h1>{e(project)}</h1><p class="lead">{e(conclusion)}</p></div><div class="meta"><strong>返点毛利与投放效果分析</strong>生成时间 {e(generated_at)}<br>数据源：{e(plan_name)} / {e(effect_name)}</div></header>
<section class="metrics">{metric_cards}</section>
<section class="grid"><article class="panel"><h2>达人表现与复投判断</h2><div class="chips">{label_chips}</div><div class="scroll"><table><thead><tr><th>达人 / 记录</th><th>判断</th><th>自然 CTR</th><th>纯 K CPC</th><th>互动量</th><th>广告占比</th><th>搜索进店成本</th></tr></thead><tbody>{ranking_rows}</tbody></table></div></article>
<aside class="panel"><h2>数据完整度</h2><h3>匹配异常</h3><ul>{manual_items}</ul><h3>关键字段缺口</h3><ul>{data_gap_items}</ul></aside></section>
<section class="panel" style="margin-top:18px"><h2>分组观察</h2><p class="method">样本少于 3 条仅作方向观察。以下结果描述相关性，不代表返点、毛利率、粉丝量级或内容类型导致效果变化。</p><div class="scroll"><table><thead><tr><th>维度</th><th>分组</th><th>样本</th><th>平均自然 CTR</th><th>平均纯 K CPC</th><th>平均搜索进店成本</th><th>结论边界</th></tr></thead><tbody>{group_rows}</tbody></table></div></section>
<section class="grid"><article class="panel"><h2>三层分析口径</h2><h3>纯 K 前端</h3><p>平均自然 CTR {e(format_number(avg_ctr, kind='percent'))}；平均纯 K CPC {e(format_number(avg_cpc, kind='currency'))}；平均互动量 {e(format_number(avg_interaction))}。</p><h3>投流潜力</h3><p>平均广告占比 {e(format_number(avg_ad_share, kind='percent'))}。占比高低只反映投入结构，必须结合推广量级和成本判断，不单独作为好坏结论。</p><h3>总投入效果</h3><p>平均搜索进店成本 {e(format_number(avg_search, kind='currency'))}。零分母或 UV 缺失的记录不参与均值。</p></article><aside class="panel"><h2>阈值与使用说明</h2><p>{e(threshold_text)}</p><p class="method">金额与业务结论保留人工复核；未定义的广告阅读成本等指标不在本报告中补造。完整字段映射、公式、源行、清洗日志与人工确认项请查看配套 Excel。</p></aside></section>
<footer>本报告与配套 Excel 由同一份清洗结果和指标口径生成。适用于内部复盘与下轮投放讨论，不替代财务核算或因果实验。</footer>
</main></body></html>"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(document, encoding="utf-8")


def analyze(args: argparse.Namespace) -> dict[str, int]:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as error:
        raise RuntimeError("缺少 openpyxl，无法读写 Excel；请使用桌面应用打包运行时") from error

    input_path = Path(args.input)
    if not input_path.exists():
        raise ValueError("项目数据文件不存在")
    source = load_workbook(input_path, read_only=True, data_only=True)
    if args.sheet and args.sheet not in source.sheetnames:
        source.close()
        raise ValueError(f"输入工作表不存在：{args.sheet}")
    sheet = source[args.sheet] if args.sheet else source.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        first_row = next(iterator)
    except StopIteration as error:
        source.close()
        raise ValueError("输入工作表为空") from error
    headers = [text_value(value) for value in first_row]
    mapping = map_headers(headers, load_json(args.column_map))
    thresholds = load_json(args.thresholds)
    records: list[dict[str, object]] = []
    duplicate_counts: defaultdict[str, int] = defaultdict(int)
    for source_row, values in enumerate(iterator, start=2):
        row: dict[str, object] = {}
        field_conflicts: list[str] = []
        for key, indexes in mapping.items():
            candidates = [values[index] for index in indexes if index < len(values) and values[index] not in (None, "")]
            unique = list(dict.fromkeys(str(value).strip() for value in candidates))
            row[key] = candidates[0] if candidates else None
            if len(unique) > 1:
                field_conflicts.append(key)
        record_id = text_value(row.get("record_id")) or f"源行{source_row}"
        duplicate_counts[record_id] += 1
        calculated = metrics(row)
        records.append({"源行": source_row, "记录ID": record_id, "字段冲突": field_conflicts, **row, **calculated})
    source.close()

    output = Workbook()
    detail = output.active
    detail.title = "指标明细"
    detail_headers = ["源行", "记录ID", "达人", "返点档", "毛利率", "达人量级", "笔记类型", *[name for name, _ in FORMULAS], "行级标签", "风险提示"]
    append_safe(detail, detail_headers)
    risk_rows: list[list[object]] = []
    for record in records:
        calculated = {name: record[name] for name, _ in FORMULAS}
        risks = []
        if duplicate_counts[str(record["记录ID"])] > 1:
            risks.append("重复记录，待确认")
        if record["字段冲突"]:
            risks.append("字段冲突：" + "、".join(record["字段冲突"]))
        missing_metrics = [name for name, value in calculated.items() if value is None]
        if missing_metrics:
            risks.append("数据不足：" + "、".join(missing_metrics))
        label = "数据不足" if risks else row_label(calculated, thresholds)
        record["行级标签"] = label
        append_safe(detail, [
            record["源行"], record["记录ID"], text_value(record.get("creator")), text_value(record.get("rebate_band")),
            number(record.get("margin_rate")), text_value(record.get("creator_tier")), text_value(record.get("note_type")),
            *[record[name] if record[name] is not None else "数据不足" for name, _ in FORMULAS], label, "；".join(risks),
        ])
        if risks:
            risk_rows.append([record["源行"], record["记录ID"], "；".join(risks), "人工复核"])

    definitions = output.create_sheet("口径表")
    append_safe(definitions, ["指标", "公式 / 口径", "说明"])
    for name, formula in FORMULAS:
        append_safe(definitions, [name, formula, "业务计划明确口径"])
    append_safe(definitions, ["其他成本指标", "待业务方提供公式", "不得自行定义"])

    grouping = output.create_sheet("分组分析")
    append_safe(grouping, ["分组维度", "分组值", "样本数", "平均自然 CTR", "平均纯 K CPC", "说明"])
    groupable_records = [
        record
        for record in records
        if duplicate_counts[str(record["记录ID"])] == 1
        and not record["字段冲突"]
        and all(record[name] is not None for name, _ in FORMULAS)
    ]
    for field, label in (("rebate_band", "返点档"), ("margin_rate", "毛利率"), ("creator_tier", "达人量级"), ("note_type", "笔记类型")):
        groups: defaultdict[str, list[dict[str, object]]] = defaultdict(list)
        for record in groupable_records:
            groups[text_value(record.get(field)) or "未提供"].append(record)
        for group_name, items in groups.items():
            ctrs = [item["自然 CTR"] for item in items if item["自然 CTR"] is not None]
            cpcs = [item["纯 K CPC"] for item in items if item["纯 K CPC"] is not None]
            append_safe(grouping, [label, group_name, len(items), sum(ctrs) / len(ctrs) if ctrs else "数据不足", sum(cpcs) / len(cpcs) if cpcs else "数据不足", "小样本仅供观察" if len(items) < 3 else ""])

    recommendations = output.create_sheet("可复投建议")
    append_safe(recommendations, ["源行", "记录ID", "达人", "行级标签", "建议", "证据摘要", "人工复核"])
    suggestion_by_label = {
        "优先保留": "优先纳入复投候选",
        "可加测": "小预算加测后再决定",
        "谨慎": "降低优先级并复核内容与成本",
        "数据不足": "补齐数据或业务阈值后再判断",
    }
    for record in records:
        evidence = "；".join(
            f"{name}={record[name]}" if record[name] is not None else f"{name}=数据不足"
            for name in ("自然 CTR", "纯 K CPC", "搜索进店成本")
        )
        label = str(record["行级标签"])
        append_safe(
            recommendations,
            [
                record["源行"],
                record["记录ID"],
                text_value(record.get("creator")),
                label,
                suggestion_by_label[label],
                evidence,
                "是",
            ],
        )

    risks_sheet = output.create_sheet("风险提示")
    append_safe(risks_sheet, ["源行", "记录ID", "风险", "处理"])
    for risk in risk_rows:
        append_safe(risks_sheet, risk)
    if not thresholds:
        append_safe(risks_sheet, ["", "", "未提供业务阈值，行级标签均为数据不足", "补充阈值后重新分析"])

    clean_log = output.create_sheet("清洗日志")
    append_safe(clean_log, ["项目", "结果"])
    append_safe(clean_log, ["源记录数", len(records)])
    append_safe(clean_log, ["重复记录数", sum(count - 1 for count in duplicate_counts.values() if count > 1)])
    append_safe(clean_log, ["字段冲突记录数", sum(1 for record in records if record["字段冲突"])])
    append_safe(clean_log, ["进入分组分析记录数", len(groupable_records)])
    append_safe(clean_log, ["源文件", input_path.name])
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output.save(output_path)
    return {"records": len(records), "risks": len(risk_rows)}


def analyze_two_sources(args: argparse.Namespace) -> dict[str, int]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as error:
        raise RuntimeError("缺少 openpyxl，无法读写 Excel；请使用桌面应用打包运行时") from error

    records, manual_rows, method_rows, thresholds, source_info = prepare_two_source_records(args)
    output = Workbook()
    detail = output.active
    detail.title = "清洗底表"
    detail_headers = [
        "计划单源行", "效果表源行", "效果记录ID", "笔记链接", "计划单达人昵称", "效果表博主昵称",
        "返点档", "毛利率", "粉丝量级", "笔记类型", "K金额", "广告金额", "自然曝光量", "自然阅读量",
        "推广曝光量", "推广阅读量", "点赞量", "评论量", "收藏量", "搜索进店UV", "总金额", "自然 CTR",
        "纯 K CPC", "互动量", "广告占比", "搜索进店成本", "行级标签", "风险提示",
    ]
    append_safe(detail, detail_headers)
    risks: list[list[object]] = []
    for record in records:
        calculated = {name: record[name] for name, _ in FORMULAS}
        missing_metrics = [name for name, value in calculated.items() if value is None]
        risk_text = "数据不足：" + "、".join(missing_metrics) if missing_metrics else ""
        label = "数据不足" if missing_metrics else row_label(calculated, thresholds)
        record["行级标签"] = label
        if risk_text:
            risks.append([record["计划单源行"], record["效果表源行"], record["笔记链接"], risk_text, "人工复核"])
        append_safe(detail, [
            record["计划单源行"], record["效果表源行"], text_value(record.get("effect_record_id")), record["笔记链接"],
            text_value(record.get("creator")), text_value(record.get("effect_creator")), text_value(record.get("rebate_band")),
            number(record.get("margin_rate")), text_value(record.get("creator_tier")), text_value(record.get("note_type")),
            number(record.get("k_amount")), number(record.get("ad_amount")), number(record.get("natural_impressions")),
            number(record.get("natural_reads")), number(record.get("promoted_impressions")), number(record.get("promoted_reads")),
            number(record.get("likes")), number(record.get("comments")), number(record.get("favorites")), number(record.get("search_uv")),
            *[record[name] if record[name] is not None else "数据不足" for name, _ in FORMULAS], label, risk_text,
        ])

    mapping_sheet = output.create_sheet("字段映射")
    append_safe(mapping_sheet, ["逻辑字段", "计划单表头", "效果表表头", "用途"])
    mapping_rows = [
        ("唯一主键", PLAN_KEY_ALIASES, EFFECT_KEY_ALIASES, "只按该字段精确一对一匹配"),
        ("达人昵称", PLAN_FIELD_ALIASES["creator"], EFFECT_FIELD_ALIASES["effect_creator"], "仅供人工核对，不参与匹配"),
        ("K金额", PLAN_FIELD_ALIASES["k_amount"], (), "纯 K CPC 与总金额"),
        ("广告金额", (), EFFECT_FIELD_ALIASES["ad_amount"], "广告占比与总金额"),
        ("自然曝光量", (), EFFECT_FIELD_ALIASES["natural_impressions"], "自然 CTR 分母"),
        ("自然阅读量", (), EFFECT_FIELD_ALIASES["natural_reads"], "自然 CTR 分子与纯 K CPC 分母"),
        ("互动量", (), EFFECT_FIELD_ALIASES["likes"] + EFFECT_FIELD_ALIASES["comments"] + EFFECT_FIELD_ALIASES["favorites"], "点赞 + 评论 + 收藏"),
        ("搜索进店UV", (), EFFECT_FIELD_ALIASES["search_uv"], "搜索进店成本分母"),
    ]
    for logical, plan_aliases, effect_aliases, purpose in mapping_rows:
        plan_header = source_info["plan_key_header"] if logical == "唯一主键" else first_header(source_info["plan_headers"], plan_aliases)
        effect_header = source_info["effect_key_header"] if logical == "唯一主键" else first_header(source_info["effect_headers"], effect_aliases)
        append_safe(mapping_sheet, [logical, plan_header or "不适用/未发现", effect_header or "不适用/未发现", purpose])

    grouping = output.create_sheet("分组分析")
    append_safe(grouping, ["分组维度", "分组值", "样本数", "平均自然 CTR", "平均纯 K CPC", "平均搜索进店成本", "说明"])
    valid_records = [record for record in records if record["行级标签"] != "数据不足"]
    for field, label in (("rebate_band", "返点档"), ("margin_rate", "毛利率"), ("creator_tier", "粉丝量级"), ("note_type", "笔记类型")):
        groups: defaultdict[str, list[dict[str, object]]] = defaultdict(list)
        for record in valid_records:
            groups[text_value(record.get(field)) or "未提供"].append(record)
        for group_name, items in groups.items():
            append_safe(grouping, [
                label, group_name, len(items),
                sum(float(item["自然 CTR"]) for item in items) / len(items),
                sum(float(item["纯 K CPC"]) for item in items) / len(items),
                sum(float(item["搜索进店成本"]) for item in items) / len(items),
                "小样本，仅作相关性观察" if len(items) < 3 else "相关性观察，不代表因果",
            ])

    conclusions = output.create_sheet("分析结论")
    append_safe(conclusions, ["笔记链接", "达人", "行级标签", "三项证据", "建议", "人工复核"])
    suggestion_by_label = {
        "优先保留": "优先纳入复投候选",
        "可加测": "小预算加测后再决定",
        "谨慎": "降低优先级并复核内容与成本",
        "数据不足": "补齐数据后再判断",
    }
    for record in records:
        label = str(record["行级标签"])
        evidence = "；".join(
            f"{name}={record[name]}" if record[name] is not None else f"{name}=数据不足"
            for name in ("自然 CTR", "纯 K CPC", "搜索进店成本")
        )
        append_safe(conclusions, [record["笔记链接"], text_value(record.get("creator")), label, evidence, suggestion_by_label[label], "是"])

    definitions = output.create_sheet("口径表")
    if method_rows:
        for row in method_rows:
            append_safe(definitions, row)
    else:
        append_safe(definitions, ["指标", "公式 / 口径", "说明"])
        for name, formula in FORMULAS:
            append_safe(definitions, [name, formula, "业务材料未提供阈值时不下行级结论"])
    append_safe(definitions, ["边界", "未定义的广告阅读成本等指标", "列为口径缺失，不自行补公式"])

    manual = output.create_sheet("人工确认")
    append_safe(manual, ["来源文件", "源行", "笔记链接", "原因", "处理状态"])
    for row in manual_rows:
        append_safe(manual, row)
    for row in risks:
        append_safe(manual, [f"{source_info['plan_path'].name} + {source_info['effect_path'].name}", f"计划单 {row[0]} / 效果表 {row[1]}", row[2], row[3], row[4]])

    clean_log = output.create_sheet("清洗日志")
    append_safe(clean_log, ["项目", "结果"])
    append_safe(clean_log, ["计划单", source_info["plan_path"].name])
    append_safe(clean_log, ["计划单工作表", source_info["plan_sheet"]])
    append_safe(clean_log, ["效果表", source_info["effect_path"].name])
    append_safe(clean_log, ["效果表工作表", source_info["effect_sheet"]])
    append_safe(clean_log, ["唯一匹配键", f"{source_info['plan_key_header']} ↔ {source_info['effect_key_header']}（精确匹配）"])
    append_safe(clean_log, ["计划单记录数", source_info["plan_count"]])
    append_safe(clean_log, ["效果表记录数", source_info["effect_count"]])
    append_safe(clean_log, ["一对一匹配记录数", len(records)])
    append_safe(clean_log, ["人工确认记录数", len(manual_rows) + len(risks)])
    append_safe(clean_log, ["阈值来源", "计划单口径 Sheet + 用户显式 thresholds（后者优先）"])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for sheet in output.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column in sheet.columns:
            letter = column[0].column_letter
            max_length = min(max((len(str(cell.value)) if cell.value is not None else 0 for cell in column), default=0) + 2, 42)
            sheet.column_dimensions[letter].width = max(10, max_length)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output.save(output_path)
    if args.html_output:
        build_html_report(
            Path(args.html_output),
            records,
            grouped_summaries(records),
            manual_rows + risks,
            source_info,
            thresholds,
        )
    return {"records": len(records), "manual_review": len(manual_rows) + len(risks)}


def self_test() -> dict[str, object]:
    values = metrics({"k_amount": 1000, "ad_amount": 200, "natural_reads": 100, "natural_impressions": 1000, "likes": 10, "comments": 5, "favorites": 5, "search_uv": 0})
    return {
        "total_amount": values["总金额"],
        "natural_ctr": values["自然 CTR"],
        "pure_k_cpc": values["纯 K CPC"],
        "zero_denominator": "数据不足" if values["搜索进店成本"] is None else values["搜索进店成本"],
        "no_threshold_label": row_label(values, None),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="分析达人项目毛利与效果")
    parser.add_argument("--input")
    parser.add_argument("--plan-input")
    parser.add_argument("--effect-input")
    parser.add_argument("--output")
    parser.add_argument("--html-output")
    parser.add_argument("--sheet")
    parser.add_argument("--plan-sheet")
    parser.add_argument("--effect-sheet")
    parser.add_argument("--method-sheet")
    parser.add_argument("--column-map")
    parser.add_argument("--thresholds")
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args()
    if args.self_test:
        print(json.dumps(self_test(), ensure_ascii=False))
        return 0
    two_source_mode = bool(args.plan_input or args.effect_input)
    if not args.output or (two_source_mode and not (args.plan_input and args.effect_input)) or (not two_source_mode and not args.input):
        print("单表模式需提供 --input；双表模式需同时提供 --plan-input 和 --effect-input；两种模式都需提供 --output", file=sys.stderr)
        return 2
    if args.html_output and not two_source_mode:
        print("HTML 分析报告当前仅支持双表模式", file=sys.stderr)
        return 2
    try:
        result = analyze_two_sources(args) if two_source_mode else analyze(args)
    except (OSError, ValueError, RuntimeError, StopIteration) as error:
        print(f"分析失败：{error}", file=sys.stderr)
        return 1
    print(json.dumps(result, ensure_ascii=False))
    print(f"ONMYAGENT_DELIVERABLE: {args.output}")
    if args.html_output:
        print(f"ONMYAGENT_DELIVERABLE: {args.html_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
