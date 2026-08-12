#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Rebate Contract Checker - Core reconciliation script
Usage:
    python check_rebate_contracts.py \
        --invoice-folder /path/to/invoices/ \
        --contract-folder /path/to/contracts/ \
        --plan-folder /path/to/plans/ \
        --output /path/to/output.xlsx

Or as a module:
    from check_rebate_contracts import RebateContractChecker
    checker = RebateContractChecker()
    results = checker.batch_check(invoice_folder, contract_folder, plan_folder)
    checker.save_results(results, output_path)
"""
import re
import os
import glob
import argparse
import json
import sys
from collections import defaultdict


class RebateContractChecker:
    """
    Cross-checks invoice applications, rebate contracts, and project plans.
    """

    # Default column aliases for invoice Excel (logical field → possible headers).
    # First match in the workbook header row wins; override with --column-map.
    DEFAULT_COLUMN_MAP = {
        '项目': ['店铺名称（必填）', '项目', '项目名称', '合作项目'],
        '服务项目名称': ['服务项目名称（必填）', '服务项目名称', '服务项目'],
        '发票抬头': ['发票抬头', '公司名称', '甲方名称', '主体名称', '对公主体'],
        '税号': ['税号', '统一社会信用代码', '纳税人识别号'],
        '开票金额': ['开票金额', '金额', '价税合计'],
        '发票类型': ['发票类型', '开票类型'],
        '开票内容': ['开票内容', '发票内容', '开票项目'],
        '地址电话': ['地址电话', '地址及电话', '注册地址'],
        '开户行及账号': ['开户行及账号', '开户行', '银行账号'],
    }

    # Preferred plan sheet name (hint only; missing → scan other sheets)
    DEFAULT_PLAN_SHEET = '推广计划单-立项及结案PM更新'

    def __init__(self, column_map=None, plan_sheet=None, amount_tolerance=1):
        self.column_map = self._normalise_column_map(column_map or self.DEFAULT_COLUMN_MAP)
        self.plan_sheet = plan_sheet or self.DEFAULT_PLAN_SHEET
        self.amount_tolerance = float(amount_tolerance)

    @staticmethod
    def _normalise_column_map(column_map):
        normalised = {}
        for field, source in (column_map or {}).items():
            if isinstance(source, (list, tuple)):
                aliases = [str(item).strip() for item in source if str(item).strip()]
            else:
                aliases = [str(source).strip()] if str(source).strip() else []
            if aliases:
                normalised[str(field)] = aliases
        return normalised

    @staticmethod
    def _header_index(headers, aliases):
        for alias in aliases:
            if alias in headers:
                return headers.index(alias)
        return None

    @staticmethod
    def _load_openpyxl():
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError as error:
            raise RuntimeError('缺少 openpyxl，无法读写 Excel；请使用桌面应用打包运行时') from error
        return Workbook, load_workbook

    @staticmethod
    def _number(value):
        if value in (None, ''):
            return 0
        try:
            return float(str(value).replace(',', '').replace('¥', '').replace('￥', '').strip())
        except (TypeError, ValueError):
            return 0

    def parse_invoice_excel(self, filepath):
        """Parse invoice application Excel and extract records."""
        _, load_workbook = self._load_openpyxl()
        workbook = load_workbook(filepath, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            first_row = next(rows)
        except StopIteration:
            workbook.close()
            raise ValueError('开票申请工作表为空')
        headers = [str(value or '').strip() for value in first_row]
        indexes = {
            field: self._header_index(headers, aliases)
            for field, aliases in self.column_map.items()
        }
        indexes = {field: index for field, index in indexes.items() if index is not None}
        records = []
        for row in rows:
            def value(field):
                index = indexes.get(field)
                return row[index] if index is not None and index < len(row) else None

            if value('发票抬头') in (None, ''):
                continue
            invoice_title = str(value('发票抬头') or '').replace('\n', ' ').strip()
            invoice_title = re.sub(r'\s*税\s*号\s*[:：].*$', '', invoice_title).strip()
            tax_id = str(value('税号')).replace('\n', ' ').strip() if value('税号') not in (None, '') else ''
            records.append({
                '来源文件': os.path.basename(filepath),
                '项目': str(value('项目') or '').strip(),
                '服务项目名称': str(value('服务项目名称') or '').strip(),
                '发票抬头': invoice_title,
                '税号': tax_id,
                '开票金额': self._number(value('开票金额')),
                '发票类型': str(value('发票类型') or '').strip(),
                '开票内容': str(value('开票内容') or '').strip(),
                '地址电话': str(value('地址电话')).replace('\n', ' ').strip() if value('地址电话') not in (None, '') else '',
                '开户行及账号': str(value('开户行及账号')).replace('\n', ' ').strip() if value('开户行及账号') not in (None, '') else '',
            })
        workbook.close()
        return records

    def parse_pdf_contract(self, filepath):
        """Parse rebate contract PDF and extract key info."""
        try:
            import pdfplumber
        except ImportError as error:
            raise RuntimeError('缺少 pdfplumber，无法读取 PDF；请使用桌面应用打包运行时') from error
        data = {
            '来源文件': os.path.basename(filepath),
            '甲方名称': '', '税号': '', '地址': '', '开户行': '', '银行账号': '',
            '渠道服务费金额': 0, '发票类型': '', '开票内容': '', '博主列表': []
        }
        with pdfplumber.open(filepath) as pdf:
            full_text = ''
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + '\n'

        match = re.search(r'(?:甲方|委托方|乙方)?[（(]?名称[）)]?[：:]\s*([^\n]+)', full_text)
        if not match:
            match = re.search(r'甲方[：:]\s*([^\n]+)', full_text)
        if not match:
            match = re.search(r'公司名称[：:]\s*([^\n]+)', full_text)
        if match:
            data['甲方名称'] = match.group(1).strip()

        match = re.search(r'税\s*号[：:]\s*([A-Z0-9]+)', full_text)
        if not match:
            match = re.search(r'纳税人识别号[：:]\s*([A-Z0-9]+)', full_text)
        if not match:
            match = re.search(r'统一社会信用代码[：:]\s*([A-Z0-9]+)', full_text)
        if match:
            data['税号'] = match.group(1).strip()

        match = re.search(r'联系地址[：:]\s*([^\n]+)', full_text)
        if match:
            data['地址'] = match.group(1).strip()

        match = re.search(r'支付全款[：:]?\s*¥\s*([\d,]+\.?\d*)', full_text)
        if not match:
            match = re.search(r'渠道服务费金额.*?¥\s*([\d,]+\.?\d*)', full_text)
        if match:
            data['渠道服务费金额'] = float(match.group(1).replace(',', ''))

        if '增值税普通发票' in full_text or '普通发票' in full_text:
            data['发票类型'] = '普通发票'
        elif '增值税专用发票' in full_text or '专用发票' in full_text:
            data['发票类型'] = '专用发票'

        match = re.search(r'开票内容[：:]\s*【([^】]+)】', full_text)
        if not match:
            match = re.search(r'开票内容[：:]\s*([^\n]+)', full_text)
        if match:
            content = match.group(1).strip()
            content = content.replace('生产生\n1/2\n活服务', '生产生活服务').replace('\n', ' ')
            data['开票内容'] = content

        match = re.search(r'开户银行[：:]\s*([^\n]+)', full_text)
        if match:
            data['开户行'] = match.group(1).strip()

        match = re.search(r'银行账号[：:]\s*([\d\s]+)', full_text)
        if match:
            data['银行账号'] = match.group(1).strip().replace(' ', '').replace('\n', '')

        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    header = [str(c).strip() if c else '' for c in table[0]]
                    header_str = ' '.join(header)
                    if '账号' in header_str and ('渠道服务费' in header_str or '合作金额' in header_str):
                        for row in table[1:]:
                            if len(row) >= 2 and row[1] and str(row[1]).strip():
                                blogger_name = str(row[1]).strip()
                                coop_amount = 0
                                fee_amount = 0
                                for cell in row:
                                    if cell and '¥' in str(cell):
                                        amounts = re.findall(r'¥\s*([\d,]+\.?\d*)', str(cell))
                                        for amt in amounts:
                                            val = float(amt.replace(',', ''))
                                            if coop_amount == 0:
                                                coop_amount = val
                                            elif fee_amount == 0:
                                                fee_amount = val
                                data['博主列表'].append({
                                    '博主名称': blogger_name,
                                    '合作金额': coop_amount,
                                    '渠道服务费': fee_amount
                                })
        return data

    def parse_plan_settlement(self, filepath):
        """
        Parse settlement-area bloggers from a project plan Excel.
        Preferred sheet is a hint; if missing or empty, scan other sheets for
        the same anchors (支出小计 … 返点小计). Layout columns are heuristic —
        override upstream materials or use a one-off helper when layout is too irregular.
        """
        _, load_workbook = self._load_openpyxl()
        workbook = load_workbook(filepath, read_only=True, data_only=True)
        sheet_order = []
        if self.plan_sheet in workbook.sheetnames:
            sheet_order.append(self.plan_sheet)
        sheet_order.extend(
            name for name in workbook.sheetnames if name not in sheet_order
        )

        bloggers = []
        for sheet_name in sheet_order:
            values = list(workbook[sheet_name].iter_rows(values_only=True))
            parsed = self._extract_plan_bloggers(values, filepath)
            if parsed:
                bloggers = parsed
                break
        workbook.close()
        return bloggers

    def _extract_plan_bloggers(self, values, filepath):
        start_row = None
        for i, row in enumerate(values):
            for val in row[:20]:
                if isinstance(val, str) and ('支出小计' in val or '达人结算' in val or '返点明细' in val):
                    start_row = i + 1
                    break
            if start_row is not None:
                break

        if start_row is None:
            return []

        end_row = len(values)
        for i in range(start_row, len(values)):
            for val in values[i][:20]:
                if isinstance(val, str) and (
                    '返点小计' in val or '支出合计' in val or '最终结算收入' in val or '合计' == val.strip()
                ):
                    end_row = i
                    break
            if end_row < len(values):
                break

        bloggers = []
        for i in range(start_row, end_row):
            row = values[i]
            blogger_name = row[15] if len(row) > 15 else None
            rebate_amount = row[16] if len(row) > 16 else None
            payment_type = row[14] if len(row) > 14 else None
            date_val = row[13] if len(row) > 13 else None
            rebate_amount_2 = row[18] if len(row) > 18 else None

            # Fallback: first non-empty text-ish cell as name, last number-ish as amount
            if blogger_name in (None, '') and row:
                for cell in row:
                    text = str(cell).strip() if cell not in (None, '') else ''
                    if text and not re.match(r'^[\d.,¥￥]+$', text) and '小计' not in text:
                        blogger_name = text
                        break

            if blogger_name not in (None, '') and str(blogger_name).strip():
                name = str(blogger_name).strip()
                if '小计' in name or '合计' in name or '最终' in name or '支出' in name:
                    continue

                amount = self._number(rebate_amount)
                if amount == 0 and rebate_amount_2 not in (None, ''):
                    amount = self._number(rebate_amount_2)

                bloggers.append({
                    '博主名称': name,
                    '计划单返点金额': round(amount, 2),
                    '付款方式': str(payment_type).strip() if payment_type not in (None, '') else '',
                    '日期': str(date_val).strip() if date_val not in (None, '') else '',
                    '来源文件': os.path.basename(filepath),
                })
        return bloggers

    def match_invoice_to_contract(self, invoice_records, contracts):
        """Match invoice records to contracts and prepare result rows."""
        results = []
        for inv in invoice_records:
            exact = [contract for contract in contracts.values() if inv['发票抬头'] == contract['甲方名称']]
            candidates = exact
            match_note = ''
            if len(exact) > 1:
                match_note = f'多协议候选（{len(exact)}份），需人工确认'
            elif not exact:
                invoice_name = self._normalise_company(inv['发票抬头'])
                candidates = [
                    contract for contract in contracts.values()
                    if self._normalise_company(contract['甲方名称']) == invoice_name
                    or (invoice_name and invoice_name in self._normalise_company(contract['甲方名称']))
                    or (self._normalise_company(contract['甲方名称']) and self._normalise_company(contract['甲方名称']) in invoice_name)
                ]
                if candidates:
                    match_note = f'模糊匹配候选（{len(candidates)}份），需人工确认'

            if not candidates:
                results.append(self._create_unmatched_row(inv))
                continue

            for contract in candidates:
                amount_match = self._compare_amount(inv['开票金额'], contract['渠道服务费金额'])
                type_match = self._compare_text(inv['发票类型'], contract['发票类型'])
                tax_match = self._compare_text(inv['税号'], contract['税号'])
                bloggers = contract['博主列表'] or [None]
                for blogger in bloggers:
                    results.append(self._create_result_row(
                        inv, contract, blogger, amount_match, type_match, tax_match, match_note
                    ))
        return results

    @staticmethod
    def _normalise_company(value):
        return re.sub(r'[\s（）()·,，。.-]+', '', str(value or '')).lower()

    @staticmethod
    def _compare_text(left, right):
        if not str(left or '').strip() or not str(right or '').strip():
            return '缺失'
        return '是' if str(left).strip() == str(right).strip() else '否'

    def _compare_amount(self, left, right):
        try:
            left_value = float(left)
            right_value = float(right)
        except (TypeError, ValueError):
            return '缺失'
        if left_value == 0 or right_value == 0:
            return '缺失'
        return '是' if abs(left_value - right_value) <= self.amount_tolerance else '否'

    def _create_result_row(self, inv, contract, blogger, amount_match, type_match, tax_match, match_note=''):
        """Create a single result row for a matched invoice-contract pair."""
        return {
            '开票来源文件': inv['来源文件'],
            '所属项目': inv['项目'],
            '服务项目名称': inv['服务项目名称'],
            '发票抬头': inv['发票抬头'],
            '税号': inv['税号'],
            '开票金额': inv['开票金额'],
            '发票类型': inv['发票类型'],
            '开票内容': inv['开票内容'],
            '博主名称': blogger['博主名称'] if blogger else 'N/A',
            '返点协议合作金额': blogger['合作金额'] if blogger else 0,
            '返点协议渠道服务费': blogger['渠道服务费'] if blogger else 0,
            '对应返点协议': contract['来源文件'],
            '协议甲方名称': contract['甲方名称'],
            '协议税号': contract['税号'],
            '协议金额': contract['渠道服务费金额'],
            '协议发票类型': contract['发票类型'],
            '协议开票内容': contract['开票内容'],
            '金额是否一致': amount_match,
            '发票类型是否一致': type_match,
            '税号是否一致': tax_match,
            '开票内容是否一致': self._compare_text(inv['开票内容'], contract['开票内容']),
            '协议匹配标记': match_note,
            '协议达人明细标记': '' if blogger else '协议中未提取到达人返点明细，三方核对不完整',
            '计划单返点金额': 0,
            '计划单来源文件': '',
            '计划单付款方式': '',
            '计划单日期': '',
            '博主金额是否一致': 'N/A',
            '重复博主标记': '',
            '核对结果': '待核对',
        }

    def _create_unmatched_row(self, inv):
        """Create a result row for an invoice with no matching contract."""
        return {
            '开票来源文件': inv['来源文件'],
            '所属项目': inv['项目'],
            '服务项目名称': inv['服务项目名称'],
            '发票抬头': inv['发票抬头'],
            '税号': inv['税号'],
            '开票金额': inv['开票金额'],
            '发票类型': inv['发票类型'],
            '开票内容': inv['开票内容'],
            '博主名称': 'N/A',
            '返点协议合作金额': 0,
            '返点协议渠道服务费': 0,
            '对应返点协议': '未找到',
            '协议甲方名称': 'N/A',
            '协议税号': 'N/A',
            '协议金额': 0,
            '协议发票类型': 'N/A',
            '协议开票内容': 'N/A',
            '金额是否一致': 'N/A',
            '发票类型是否一致': 'N/A',
            '税号是否一致': 'N/A',
            '开票内容是否一致': 'N/A',
            '协议匹配标记': '',
            '协议达人明细标记': '未找到协议，三方核对不完整',
            '计划单返点金额': 0,
            '计划单来源文件': '',
            '计划单付款方式': '',
            '计划单日期': '',
            '博主金额是否一致': 'N/A',
            '重复博主标记': '',
            '核对结果': '未找到对应返点协议',
        }

    def enrich_with_plan_data(self, results, plan_blogger_data):
        """Enrich unique plan matches; duplicate bloggers always require manual review."""
        blogger_plan_count = defaultdict(list)
        for _plan_file, bloggers in plan_blogger_data.items():
            for b in bloggers:
                blogger_plan_count[b['博主名称']].append(b)

        for r in results:
            blogger_name = r['博主名称']
            if blogger_name == 'N/A':
                continue
            matches = blogger_plan_count.get(blogger_name, [])
            if not matches:
                normalised_name = self._normalise_company(blogger_name)
                fuzzy_matches = [
                    item
                    for candidate_name, items in blogger_plan_count.items()
                    if self._normalise_company(candidate_name) == normalised_name
                    or (normalised_name and normalised_name in self._normalise_company(candidate_name))
                    or (self._normalise_company(candidate_name) and self._normalise_company(candidate_name) in normalised_name)
                    for item in items
                ]
                if fuzzy_matches:
                    r['计划单来源文件'] = '；'.join(item['来源文件'] for item in fuzzy_matches)
                    r['博主金额是否一致'] = '需人工确认（达人名称模糊匹配）'
                    r['重复博主标记'] = f'达人名称存在{len(fuzzy_matches)}条模糊候选，不默认取第一条'
                    continue
                r['博主金额是否一致'] = 'N/A（计划单中未找到该博主）'
                continue
            if len(matches) > 1:
                plan_names = [item['来源文件'] for item in matches]
                amounts = [item['计划单返点金额'] for item in matches]
                r['计划单来源文件'] = '；'.join(plan_names)
                r['博主金额是否一致'] = '需人工确认（重复达人）'
                r['重复博主标记'] = f'出现在{len(matches)}条计划记录中：{plan_names}；金额：{amounts}'
                continue

            match = matches[0]
            r['计划单返点金额'] = match['计划单返点金额']
            r['计划单来源文件'] = match['来源文件']
            r['计划单付款方式'] = match['付款方式']
            r['计划单日期'] = match['日期']
            comparison = self._compare_amount(r['返点协议渠道服务费'], match['计划单返点金额'])
            if comparison == '否':
                diff = round(r['返点协议渠道服务费'] - match['计划单返点金额'], 2)
                r['博主金额是否一致'] = f'否（差异{diff}）'
            elif comparison == '缺失':
                r['博主金额是否一致'] = 'N/A（协议或计划单金额为空）'
            else:
                r['博主金额是否一致'] = '是'

        return results

    def determine_final_result(self, r):
        """Determine the final verification result for a row."""
        if r['对应返点协议'] == '未找到':
            return '⚠️ 未找到对应返点协议'
        manual = [
            value
            for value in (
                r.get('协议匹配标记'),
                r.get('协议达人明细标记'),
                r.get('重复博主标记'),
            )
            if value
        ]
        if manual:
            return '⚠️ 需人工确认：' + '；'.join(manual)

        checks = []
        if r['金额是否一致'] == '否':
            checks.append('开票金额不一致')
        if r['发票类型是否一致'] == '否':
            checks.append('发票类型不一致')
        if r['税号是否一致'] == '否':
            checks.append('税号不一致')
        if r['开票内容是否一致'] == '否':
            checks.append('开票内容不一致')
        for field, label in (
            ('金额是否一致', '开票金额缺失'),
            ('发票类型是否一致', '发票类型缺失'),
            ('税号是否一致', '税号缺失'),
            ('开票内容是否一致', '开票内容缺失'),
        ):
            if r.get(field) == '缺失':
                checks.append(label)
        if r['博主金额是否一致'] and r['博主金额是否一致'].startswith('否'):
            checks.append('博主返点金额与协议不一致')
        if r['博主金额是否一致'] and '未找到' in r['博主金额是否一致']:
            return '⚠️ 需人工确认：计划单中未找到该博主，三方核对不完整'
        if r['博主金额是否一致'] and '为空' in r['博主金额是否一致']:
            checks.append('金额数据缺失')
        if r['博主金额是否一致'] and '需人工确认' in r['博主金额是否一致']:
            return '⚠️ 需人工确认：重复达人计划记录'

        if not checks:
            return '✅ 核对通过'
        else:
            return '❌ ' + '；'.join(checks)

    def check(self, invoice_files=None, contract_files=None, plan_files=None):
        """
        Run the full reconciliation check on specified files.

        Args:
            invoice_files: List of paths to invoice Excel files
            contract_files: List of paths to contract PDF files
            plan_files: List of paths to plan Excel files

        Returns:
            List of result dictionaries
        """
        # Parse all documents
        all_invoice_records = []
        if invoice_files:
            for f in invoice_files:
                all_invoice_records.extend(self.parse_invoice_excel(f))

        all_contracts = {}
        if contract_files:
            for f in contract_files:
                contract = self.parse_pdf_contract(f)
                all_contracts[os.path.basename(f)] = contract

        all_plan_data = {}
        if plan_files:
            for f in plan_files:
                all_plan_data[os.path.basename(f)] = self.parse_plan_settlement(f)

        # Match and enrich
        results = self.match_invoice_to_contract(all_invoice_records, all_contracts)
        results = self.enrich_with_plan_data(results, all_plan_data)

        for r in results:
            r['核对结果'] = self.determine_final_result(r)

        return results

    def batch_check(self, invoice_folder=None, contract_folder=None, plan_folder=None):
        """
        Batch process all files in specified folders.

        Args:
            invoice_folder: Path to folder containing invoice Excel files
            contract_folder: Path to folder containing contract PDF files
            plan_folder: Path to folder containing plan Excel files

        Returns:
            List of result dictionaries
        """
        invoice_files = glob.glob(os.path.join(invoice_folder, '*.xlsx')) if invoice_folder else []
        contract_files = glob.glob(os.path.join(contract_folder, '*.pdf')) if contract_folder else []
        plan_files = glob.glob(os.path.join(plan_folder, '*.xlsx')) if plan_folder else []
        return self.check(invoice_files=invoice_files, contract_files=contract_files, plan_files=plan_files)

    def save_results(self, results, output_path):
        """
        Save results to three sheets: 核对清单, 核对摘要, and 待处理事项.

        Args:
            results: List of result dictionaries from check() or batch_check()
            output_path: Path to output Excel file
        """
        Workbook, _ = self._load_openpyxl()
        output_cols = [
            '开票来源文件', '所属项目', '服务项目名称', '发票抬头', '税号', '开票金额', '发票类型', '开票内容',
            '博主名称', '返点协议合作金额', '返点协议渠道服务费',
            '对应返点协议', '协议甲方名称', '协议税号', '协议金额', '协议发票类型', '协议开票内容',
            '金额是否一致', '发票类型是否一致', '税号是否一致', '开票内容是否一致', '协议匹配标记',
            '协议达人明细标记',
            '计划单返点金额', '计划单来源文件', '计划单付款方式', '计划单日期',
            '博主金额是否一致', '重复博主标记', '核对结果'
        ]
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
        workbook = Workbook()
        checklist = workbook.active
        checklist.title = '核对清单'
        checklist.append([self._excel_safe(value) for value in output_cols])
        for result in results:
            checklist.append([self._excel_safe(result.get(column, '')) for column in output_cols])

        summary = workbook.create_sheet('核对摘要')
        summary.append(['指标', '数量'])
        for metric, value in (
            ('金额容差（元）', self.amount_tolerance),
            ('总记录数', len(results)),
            ('核对通过', sum(1 for r in results if r['核对结果'].startswith('✅'))),
            ('核对不通过', sum(1 for r in results if r['核对结果'].startswith('❌'))),
            ('需人工确认/未找到', sum(1 for r in results if r['核对结果'].startswith('⚠️'))),
        ):
            summary.append([self._excel_safe(metric), self._excel_safe(value)])

        pending = workbook.create_sheet('待处理事项')
        pending.append([self._excel_safe(value) for value in output_cols])
        for result in results:
            if not result.get('核对结果', '').startswith('✅'):
                pending.append([self._excel_safe(result.get(column, '')) for column in output_cols])
        workbook.save(output_path)

        print(f"Results saved to: {output_path}")
        return output_path

    @staticmethod
    def _excel_safe(value):
        if isinstance(value, str) and value.startswith(('=', '+', '-', '@')):
            return "'" + value
        return value


def self_test():
    checker = RebateContractChecker(amount_tolerance=1)
    invoice = {
        '来源文件': 'invoice.xlsx', '项目': '项目A', '服务项目名称': '服务',
        '发票抬头': '示例公司', '税号': 'TAX', '开票金额': 1000,
        '发票类型': '普通发票', '开票内容': '信息服务费',
    }
    contract = {
        '来源文件': 'contract.pdf', '甲方名称': '示例公司', '税号': 'TAX',
        '渠道服务费金额': 1001, '发票类型': '普通发票', '开票内容': '信息服务费',
        '博主列表': [{'博主名称': '达人A', '合作金额': 10000, '渠道服务费': 1000}],
    }
    unique = checker.match_invoice_to_contract([invoice], {'contract.pdf': contract})[0]
    checker.enrich_with_plan_data([unique], {
        'plan.xlsx': [{
            '博主名称': '达人A', '计划单返点金额': 1000, '来源文件': 'plan.xlsx',
            '付款方式': '银行转账', '日期': '2026-08-10',
        }],
    })
    duplicate = checker.match_invoice_to_contract(
        [invoice], {'a.pdf': dict(contract, 来源文件='a.pdf'), 'b.pdf': dict(contract, 来源文件='b.pdf')}
    )[0]
    return {
        'amount_tolerance': checker.amount_tolerance,
        'duplicate_status': '需人工确认' if '需人工确认' in checker.determine_final_result(duplicate) else checker.determine_final_result(duplicate),
        'unique_status': '核对通过' if checker.determine_final_result(unique).startswith('✅') else checker.determine_final_result(unique),
        'missing_third_leg_status': checker.determine_final_result(
            checker._create_result_row(invoice, dict(contract, 博主列表=[]), None, '是', '是', '是')
        ),
    }


def load_column_map(value):
    if not value:
        return None
    if value.lstrip().startswith('{'):
        content = value
    elif os.path.exists(value):
        with open(value, encoding='utf-8') as source:
            content = source.read()
    else:
        content = value
    parsed = json.loads(content)
    if not isinstance(parsed, dict):
        raise ValueError('列映射必须是 JSON 对象')
    # Values may be a single header string or a list of header aliases
    normalised = {}
    for key, source in parsed.items():
        if isinstance(source, list):
            normalised[str(key)] = [str(item) for item in source]
        else:
            normalised[str(key)] = str(source)
    return normalised


def main():
    parser = argparse.ArgumentParser(description='Rebate Contract Checker')
    parser.add_argument('--invoice-folder', help='Folder containing invoice Excel files')
    parser.add_argument('--contract-folder', help='Folder containing contract PDF files')
    parser.add_argument('--plan-folder', help='Folder containing plan Excel files')
    parser.add_argument('--output', '-o', default='返点协议核对清单.xlsx', help='Output Excel file path')
    parser.add_argument(
        '--plan-sheet',
        default='推广计划单-立项及结案PM更新',
        help='Preferred plan sheet name (hint; other sheets scanned if missing/empty)',
    )
    parser.add_argument('--amount-tolerance', type=float, default=1, help='Amount tolerance in CNY (default: 1)')
    parser.add_argument(
        '--column-map',
        help='Invoice column mapping JSON: logical field → header string or alias list',
    )
    parser.add_argument('--self-test', action='store_true')
    args = parser.parse_args()

    if args.self_test:
        print(json.dumps(self_test(), ensure_ascii=False))
        return 0

    if not (args.invoice_folder and args.contract_folder and args.plan_folder):
        print("Usage: python check_rebate_contracts.py --invoice-folder <path> --contract-folder <path> --plan-folder <path> --output <path>")
        return 2

    try:
        checker = RebateContractChecker(
            column_map=load_column_map(args.column_map),
            plan_sheet=args.plan_sheet,
            amount_tolerance=args.amount_tolerance,
        )
        results = checker.batch_check(
            invoice_folder=args.invoice_folder,
            contract_folder=args.contract_folder,
            plan_folder=args.plan_folder
        )
        checker.save_results(results, args.output)
    except (OSError, RuntimeError, ValueError, json.JSONDecodeError) as error:
        print(f'核对失败：{error}', file=sys.stderr)
        return 1

    print(f"\nVerification Summary:")
    print(f"  Total: {len(results)}")
    print(f"  Passed: {sum(1 for r in results if r['核对结果'].startswith('✅'))}")
    print(f"  Failed: {sum(1 for r in results if r['核对结果'].startswith('❌'))}")
    print(f"  Not Found: {sum(1 for r in results if r['核对结果'].startswith('⚠️'))}")
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
