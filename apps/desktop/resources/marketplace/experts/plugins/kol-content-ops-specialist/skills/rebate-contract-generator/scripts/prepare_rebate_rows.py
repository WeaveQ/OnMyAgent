#!/usr/bin/env python3
"""Prepare per-contract JSON rows for officecli merge (no DOCX writing).

Reads rebate Excel, optionally parses free-text company blocks, multi-talent
names, amount-to-Chinese, and output filenames. Writes rows.jsonl only.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "seq": ("序号", "seq", "编号"),
    "blogger": ("博主名称", "达人", "达人名称", "博主"),
    "project": ("项目", "项目名称", "合作品牌", "品牌"),
    "signing": ("签署方式", "signing_method"),
    "party": ("主体名称", "对公主体", "公司名称", "甲方名称", "发票抬头"),
    "company_blob": ("公司信息", "对公信息", "主体信息"),
    "tax": ("统一社会信用代码", "税号", "纳税人识别号"),
    "bank": ("开户行", "开户银行"),
    "account": ("银行账号", "账号"),
    "address": ("注册地址", "地址", "联系地址"),
    "phone": ("电话", "公司电话", "联系电话"),
    "amount": ("返点金额", "渠道服务费", "渠道服务费金额", "金额"),
    "invoice_type": ("发票类型", "开票类型"),
    "invoice_content": ("发票内容", "开票内容"),
    "period": ("合作周期", "合同周期"),
    "link": ("发布链接", "链接"),
    "platform": ("平台",),
    "account_name": ("账号", "账号昵称"),
}


def _display(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _header_index(headers: list[str], aliases: tuple[str, ...]) -> int | None:
    for alias in aliases:
        if alias in headers:
            return headers.index(alias)
    return None


def parse_company_blob(text: str) -> dict[str, str]:
    out = {
        "税号": "",
        "开户行": "",
        "银行账号": "",
        "地址": "",
        "电话": "",
        "邮箱": "",
        "开户名称": "",
    }
    if not text:
        return out
    for raw in str(text).splitlines():
        line = raw.strip()
        if not line:
            continue
        match = re.match(r"([^\s:：]+)[\s:：]+(.+)", line)
        if not match:
            continue
        key, val = match.group(1).strip(), match.group(2).strip()
        val = re.sub(r"[【】\[\]（）()\s]+", "", val).strip() or match.group(2).strip()
        if not val:
            continue
        if any(token in key for token in ("统一社会信用代码", "纳税人识别号", "税号")):
            out["税号"] = out["税号"] or val
        elif any(token in key for token in ("开户银行", "开户行")):
            out["开户行"] = out["开户行"] or val
        elif any(token in key for token in ("银行账号", "账号")) and "开户" not in key:
            out["银行账号"] = out["银行账号"] or val
        elif any(token in key for token in ("地址", "住所")):
            out["地址"] = out["地址"] or val
        elif any(token in key for token in ("电话", "手机")):
            out["电话"] = out["电话"] or val
        elif "邮箱" in key or "邮件" in key:
            out["邮箱"] = out["邮箱"] or val
        elif any(token in key for token in ("开户名称", "户名", "名称")) and not out["开户名称"]:
            out["开户名称"] = val
    return out


_DIGITS = "零壹贰叁肆伍陆柒捌玖"
_UNITS = ("", "拾", "佰", "仟")
_SECTIONS = ("", "万", "亿")


def amount_to_chinese(amount: object) -> str:
    """Convert number-like amount to uppercase Chinese yuan."""
    text = _display(amount).replace(",", "").replace("¥", "").replace("￥", "")
    if not text:
        return ""
    try:
        value = float(text)
    except ValueError:
        return ""
    if value < 0:
        return ""
    integer = int(round(value * 100))
    if integer == 0:
        return "零元整"
    jiao = (integer // 10) % 10
    fen = integer % 10
    yuan = integer // 100

    def int_to_chinese(n: int) -> str:
        if n == 0:
            return _DIGITS[0]
        chunks: list[str] = []
        unit_index = 0
        while n > 0:
            section = n % 10000
            if section != 0:
                sec_str = ""
                tmp = section
                u = 0
                zero_pending = False
                while tmp > 0:
                    digit = tmp % 10
                    if digit == 0:
                        if sec_str and not zero_pending:
                            zero_pending = True
                            sec_str = _DIGITS[0] + sec_str
                    else:
                        zero_pending = False
                        sec_str = _DIGITS[digit] + _UNITS[u] + sec_str
                    u += 1
                    tmp //= 10
                chunks.append(sec_str + _SECTIONS[unit_index])
            elif chunks:
                chunks.append(_DIGITS[0])
            unit_index += 1
            n //= 10000
        text_out = "".join(reversed(chunks))
        while "零零" in text_out:
            text_out = text_out.replace("零零", "零")
        return text_out.strip("零") or _DIGITS[0]

    body = int_to_chinese(yuan) + "元"
    if jiao == 0 and fen == 0:
        return body + "整"
    if jiao != 0:
        body += _DIGITS[jiao] + "角"
    if fen != 0:
        body += _DIGITS[fen] + "分"
    return body


def party_short_name(name: str) -> str:
    text = re.sub(r"^(北京市|上海市|天津市|重庆市|[\u4e00-\u9fff]{2,3}省|[\u4e00-\u9fff]{2,3}市)", "", name)
    text = re.sub(
        r"(有限责任公司|股份有限公司|有限公司|集团|公司|工作室|工作室部)$",
        "",
        text,
    )
    text = text.strip()
    if len(text) <= 4:
        return text or name[:4]
    return text[:4]


def safe_filename(name: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|\n\r\t]+', "_", name)
    cleaned = cleaned.replace("/", "／").replace("／", "_")
    return cleaned.strip(" ._") or "合同"


def split_bloggers(raw: str) -> list[str]:
    if not raw:
        return []
    parts = re.split(r"[\n、,/，;；|]+", raw)
    return [part.strip() for part in parts if part.strip()]


def build_rows(input_path: Path, sheet_name: str | None) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("缺少 openpyxl") from error

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    if sheet_name and sheet_name not in workbook.sheetnames:
        raise ValueError(f"工作表不存在：{sheet_name}")
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = [_display(value) for value in next(rows_iter)]
    except StopIteration as error:
        raise ValueError("工作表为空") from error

    indexes = {
        key: _header_index(headers, aliases) for key, aliases in FIELD_ALIASES.items()
    }

    def cell(values: list[object], key: str) -> str:
        idx = indexes.get(key)
        if idx is None or idx >= len(values):
            return ""
        return _display(values[idx])

    results: list[dict[str, object]] = []
    for row_number, values in enumerate(rows_iter, start=2):
        values_list = list(values)
        if not any(_display(value) for value in values_list):
            continue
        party = cell(values_list, "party")
        blob = parse_company_blob(cell(values_list, "company_blob"))
        tax = cell(values_list, "tax") or blob["税号"]
        bank = cell(values_list, "bank") or blob["开户行"]
        account = cell(values_list, "account") or blob["银行账号"]
        address = cell(values_list, "address") or blob["地址"]
        phone = cell(values_list, "phone") or blob["电话"]
        amount = cell(values_list, "amount")
        project = cell(values_list, "project")
        signing = cell(values_list, "signing")
        bloggers = split_bloggers(cell(values_list, "blogger"))
        if not bloggers:
            bloggers = [""]

        blogger_label = "／".join(bloggers) if len(bloggers) > 1 else bloggers[0]
        project_token = re.sub(r"[\s\-]+", "", project).lower() or "项目"
        short = party_short_name(party) if party else "主体"
        output_name = safe_filename(f"【{blogger_label}-{project_token}-{signing or '签署'}-{short}】")

        merge = {
            "合同编号": "",
            "签署方式": signing,
            "甲方名称": party or blob["开户名称"],
            "甲方地址": address,
            "联系人": "",
            "联系电话": phone,
            "税号": tax,
            "乙方名称": "【待确认】",
            "乙方地址": "【待确认】",
            "乙方联系人": "【待确认】",
            "乙方电话": "【待确认】",
            "合作品牌": project,
            "渠道服务费金额": amount,
            "渠道服务费大写": amount_to_chinese(amount),
            "开票名称": party or blob["开户名称"],
            "开户行": bank,
            "银行账号": account,
            "发票类型": cell(values_list, "invoice_type") or "普通",
            "开票内容": cell(values_list, "invoice_content") or "生产生活服务*信息服务费",
            "开票地址": address,
            "开票电话": phone,
            "合作周期": cell(values_list, "period"),
        }

        table_rows = []
        for index, blogger in enumerate(bloggers):
            table_rows.append(
                {
                    "序号": str(index + 1),
                    "平台": cell(values_list, "platform"),
                    "账号": blogger or cell(values_list, "account_name"),
                    "发布链接": cell(values_list, "link"),
                    "合作金额": amount if index == 0 else "",
                    "渠道服务费": amount if index == 0 else "",
                }
            )

        missing = [
            label
            for label, key in (
                ("甲方名称", "甲方名称"),
                ("税号", "税号"),
                ("开户行", "开户行"),
                ("银行账号", "银行账号"),
                ("渠道服务费金额", "渠道服务费金额"),
            )
            if not str(merge.get(key, "")).strip()
        ]
        notes = []
        if len(bloggers) > 1:
            notes.append(f"双博主已展开{len(bloggers)}行")
        if missing:
            notes.append("缺" + "、".join(missing))

        results.append(
            {
                "source_row": row_number,
                "output_name": output_name,
                "merge": merge,
                "table_rows": table_rows,
                "missing": missing,
                "notes": notes,
                "status": "ready" if not missing else "incomplete",
            }
        )
    workbook.close()
    return results


def main() -> int:
    parser = argparse.ArgumentParser(description="Prepare rebate merge rows (JSONL only)")
    parser.add_argument("--input", help="返点信息 Excel")
    parser.add_argument("--sheet", help="工作表名")
    parser.add_argument(
        "--out-dir",
        default=".opencode/tmp",
        help="过程 JSON 输出目录（默认 .opencode/tmp，不当产物）",
    )
    parser.add_argument("--rows", default="rows.jsonl", help="JSONL 文件名")
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args()

    if args.self_test:
        assert "柒" in amount_to_chinese(720) or "柒" in amount_to_chinese("720")
        blob = parse_company_blob("税号：91310000TEST\n开户行：示例银行\n银行账号：62220000")
        assert blob["税号"].startswith("9131")
        print(json.dumps({"ok": 1}, ensure_ascii=False))
        return 0

    if not args.input:
        print("必须提供 --input", file=sys.stderr)
        return 2

    try:
        rows = build_rows(Path(args.input), args.sheet)
    except (OSError, ValueError, RuntimeError) as error:
        print(f"prepare 失败：{error}", file=sys.stderr)
        return 1

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    rows_path = out_dir / args.rows
    with rows_path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")
            # Also write per-row merge JSON for officecli --data
            if row.get("status") == "ready":
                merge_path = out_dir / f"row_{row['source_row']}.json"
                merge_path.write_text(
                    json.dumps(row["merge"], ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )

    summary = {
        "rows": len(rows),
        "ready": sum(1 for row in rows if row.get("status") == "ready"),
        "incomplete": sum(1 for row in rows if row.get("status") != "ready"),
        "rows_file": str(rows_path),
    }
    print(json.dumps(summary, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
