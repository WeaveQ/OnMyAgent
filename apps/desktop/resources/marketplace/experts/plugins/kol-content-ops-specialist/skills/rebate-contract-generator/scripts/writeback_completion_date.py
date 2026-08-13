#!/usr/bin/env python3
"""Write completion dates to a session deliverable and, when authorized, its source."""

from __future__ import annotations

import argparse
import json
import os
import re
import stat
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


SEQUENCE_HEADERS = ("序号", "编号", "seq")
COMPLETION_HEADERS = ("完成日期", "完成后填写日期", "填写完成日期")


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _column_index(value: str | None) -> int | None:
    if not value:
        return None
    token = value.strip()
    if token.isdigit():
        index = int(token)
    else:
        index = column_index_from_string(token.upper())
    if index < 1:
        raise ValueError("列号必须从 1/A 开始")
    return index


def _find_header_layout(
    sheet,
    *,
    header_row: int | None,
    sequence_column: int | None,
    completion_column: int | None,
) -> tuple[int, int, int]:
    candidate_rows = [header_row] if header_row else range(1, min(sheet.max_row, 30) + 1)
    for row_number in candidate_rows:
        if row_number is None or row_number < 1:
            continue
        values = [_text(sheet.cell(row_number, column).value) for column in range(1, sheet.max_column + 1)]
        detected_sequence = next(
            (index for index, value in enumerate(values, start=1) if value in SEQUENCE_HEADERS),
            None,
        )
        detected_completion = next(
            (index for index, value in enumerate(values, start=1) if value in COMPLETION_HEADERS),
            None,
        )
        sequence = sequence_column or detected_sequence
        completion = completion_column or detected_completion
        if sequence and completion and (header_row or detected_sequence):
            return row_number, sequence, completion
    raise ValueError("未找到序号与完成日期表头；请传 --header-row/--sequence-column/--completion-column")


def _validate_cells(path: Path, sheet_name: str, cells: list[str], value: str) -> None:
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=True)
    try:
        sheet = workbook[sheet_name]
        mismatches = [cell for cell in cells if _text(sheet[cell].value) != value]
        if mismatches:
            raise ValueError(f"回写校验失败：{', '.join(mismatches[:5])}")
    finally:
        workbook.close()


def _atomic_save(workbook, source: Path) -> None:
    original_mode = stat.S_IMODE(source.stat().st_mode)
    handle, temporary_name = tempfile.mkstemp(
        prefix=f".{source.stem}.onmyagent-",
        suffix=source.suffix,
        dir=source.parent,
    )
    os.close(handle)
    temporary = Path(temporary_name)
    try:
        workbook.save(temporary)
        os.chmod(temporary, original_mode)
        os.replace(temporary, source)
    finally:
        temporary.unlink(missing_ok=True)


def write_completion_dates(
    input_path: Path,
    output_path: Path,
    completion_date: str,
    *,
    sheet_name: str | None = None,
    header_row: int | None = None,
    sequence_column: str | None = None,
    completion_column: str | None = None,
    source_writeback: bool = False,
) -> dict[str, object]:
    input_path = input_path.resolve()
    output_path = output_path.resolve()
    if not input_path.is_file():
        raise ValueError(f"源 Excel 不存在：{input_path}")
    if input_path == output_path:
        raise ValueError("--output 必须是会话内副本，不能与 --input 相同")
    if input_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("仅支持 .xlsx/.xlsm")
    if input_path.suffix.lower() != output_path.suffix.lower():
        raise ValueError("源文件与副本扩展名必须一致")
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", completion_date):
        raise ValueError("--date 必须为 YYYY-MM-DD")

    keep_vba = input_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(input_path, read_only=False, data_only=False, keep_vba=keep_vba, keep_links=True)
    try:
        if sheet_name and sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表不存在：{sheet_name}")
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        resolved_header, resolved_sequence, resolved_completion = _find_header_layout(
            sheet,
            header_row=header_row,
            sequence_column=_column_index(sequence_column),
            completion_column=_column_index(completion_column),
        )
        cells: list[str] = []
        for row_number in range(resolved_header + 1, sheet.max_row + 1):
            if not _text(sheet.cell(row_number, resolved_sequence).value):
                continue
            cell = sheet.cell(row_number, resolved_completion)
            cell.value = completion_date
            cells.append(cell.coordinate)
        if not cells:
            raise ValueError("序号列没有可回写的数据行")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        _validate_cells(output_path, sheet.title, cells, completion_date)

        if source_writeback:
            _atomic_save(workbook, input_path)
            _validate_cells(input_path, sheet.title, cells, completion_date)
    finally:
        workbook.close()

    return {
        "status": "success",
        "deliverable": True,
        "path": str(output_path),
        "rows_updated": len(cells),
        "cells": cells,
        "source_updated": source_writeback,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="回写返点台账完成日期")
    parser.add_argument("--input", required=True, help="用户提供的源 Excel 精确路径")
    parser.add_argument("--output", required=True, help="当前专家会话内的交付副本")
    parser.add_argument("--date", required=True, help="YYYY-MM-DD")
    parser.add_argument("--sheet", help="工作表名；默认活动表")
    parser.add_argument("--header-row", type=int, help="表头行号")
    parser.add_argument("--sequence-column", help="序号列，如 A 或 1")
    parser.add_argument("--completion-column", help="完成日期列，如 B 或 2")
    parser.add_argument(
        "--source-writeback",
        action="store_true",
        help="仅在用户明确要求回填源文件/原表时使用",
    )
    args = parser.parse_args()

    try:
        result = write_completion_dates(
            Path(args.input),
            Path(args.output),
            args.date,
            sheet_name=args.sheet,
            header_row=args.header_row,
            sequence_column=args.sequence_column,
            completion_column=args.completion_column,
            source_writeback=args.source_writeback,
        )
    except (OSError, ValueError) as error:
        print(f"完成日期回写失败：{error}", file=sys.stderr)
        return 1

    print(json.dumps(result, ensure_ascii=False))
    print(f"ONMYAGENT_DELIVERABLE: {result['path']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
