#!/usr/bin/env python3
"""Batch-fill approved DOCX contract templates from Excel rows.

Primary product path is `officecli merge` (see SKILL.md). This script is the
offline / packaged-runtime fallback: same placeholder model (`{{任意名}}`),
no hard binding to English-only keys or a single Excel layout.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree


WORD_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
# Any non-empty placeholder body, including CJK (same family as officecli merge).
PLACEHOLDER_RE = re.compile(r"\{\{\s*(.+?)\s*\}\}")

# Business default required data fields (data completeness, not template shape).
DEFAULT_REQUIRED_FIELDS = (
    "counterparty_name",
    "unified_social_credit_code",
    "bank_account",
    "bank_name",
    "invoice_content",
    "invoice_type",
    "rebate_amount",
    "cooperation_period",
)

# Canonical field → Excel header aliases (hints only; user map wins).
FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "project_name": ("project_name", "项目", "项目名称", "合作项目"),
    "counterparty_name": ("counterparty_name", "主体名称", "对公主体", "公司名称", "甲方名称", "发票抬头"),
    "unified_social_credit_code": (
        "unified_social_credit_code",
        "统一社会信用代码",
        "税号",
        "纳税人识别号",
    ),
    "bank_account": ("bank_account", "银行账号", "账号"),
    "bank_name": ("bank_name", "开户行", "开户银行"),
    "registered_address": ("registered_address", "注册地址", "地址"),
    "phone": ("phone", "电话", "公司电话"),
    "contact_name": ("contact_name", "联系人", "联系人姓名"),
    "contact_phone": ("contact_phone", "联系人电话", "联系电话"),
    "signing_method": ("signing_method", "签署方式"),
    "invoice_content": ("invoice_content", "发票内容", "开票内容", "开票项目"),
    "invoice_type": ("invoice_type", "发票类型", "开票类型"),
    "rebate_amount": ("rebate_amount", "返点金额", "渠道服务费"),
    "cooperation_period": ("cooperation_period", "合作周期", "合同周期"),
    "payment_method": ("payment_method", "付款方式", "支付方式"),
}

# Reverse: any alias / canonical → canonical
_ALIAS_TO_CANONICAL: dict[str, str] = {}
for _canonical, _aliases in FIELD_ALIASES.items():
    for _alias in _aliases:
        _ALIAS_TO_CANONICAL[_alias] = _canonical
    _ALIAS_TO_CANONICAL[_canonical] = _canonical


def _display_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _load_json(value: str | None) -> dict[str, object]:
    if not value:
        return {}
    if value.lstrip().startswith("{"):
        content = value
    else:
        candidate = Path(value)
        content = candidate.read_text(encoding="utf-8") if candidate.exists() else value
    parsed = json.loads(content)
    if not isinstance(parsed, dict):
        raise ValueError("映射必须是 JSON 对象")
    return parsed


def _as_str_list(value: object) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    text = str(value).strip()
    return [text] if text else []


def _parse_map_config(raw: dict[str, object]) -> dict[str, object]:
    """Normalise --map / --column-map payloads into a single config."""
    placeholders: dict[str, str] = {}
    columns: dict[str, list[str]] = {}
    required = list(DEFAULT_REQUIRED_FIELDS)

    # New shape: { placeholders, columns|records.columns, required_fields }
    ph_src = raw.get("placeholders")
    if isinstance(ph_src, dict):
        for key, target in ph_src.items():
            placeholders[str(key).strip()] = str(target).strip()

    col_src = raw.get("columns")
    records = raw.get("records")
    if col_src is None and isinstance(records, dict):
        col_src = records.get("columns")
    if isinstance(col_src, dict):
        for key, source in col_src.items():
            columns[str(key).strip()] = _as_str_list(source)

    req = raw.get("required_fields") or raw.get("required")
    if isinstance(req, list) and req:
        required = [str(item).strip() for item in req if str(item).strip()]

    # Legacy --column-map: { canonical: header | [headers] } without wrappers
    if not placeholders and not columns:
        for key, source in raw.items():
            if key in {"version", "placeholder_pattern", "mode"}:
                continue
            columns[str(key).strip()] = _as_str_list(source)

    return {
        "placeholders": placeholders,
        "columns": columns,
        "required_fields": required,
    }


def _header_mapping(headers: list[str], custom_columns: dict[str, list[str]]) -> dict[str, list[int]]:
    aliases = {key: set(values) for key, values in FIELD_ALIASES.items()}
    for key, values in custom_columns.items():
        aliases.setdefault(key, set()).update(values)
        # Also allow mapping by excel header → keep key as-is for fill
        aliases.setdefault(key, set()).add(key)

    mapping: dict[str, list[int]] = {}
    for canonical, values in aliases.items():
        indexes = [index for index, header in enumerate(headers) if header in values]
        if indexes:
            mapping[canonical] = indexes

    # Direct header capture: every excel header is also a key for template match
    for index, header in enumerate(headers):
        if header and header not in mapping:
            mapping[header] = [index]
        elif header:
            # Ensure header name itself resolves even if already under canonical
            mapping.setdefault(header, []).append(index)
            mapping[header] = list(dict.fromkeys(mapping[header]))
    return mapping


def _normalise_record(values: list[object], mapping: dict[str, list[int]]) -> tuple[dict[str, str], list[str]]:
    record: dict[str, str] = {}
    conflicts: list[str] = []
    for key, indexes in mapping.items():
        candidates = [_display_value(values[index]) for index in indexes if index < len(values)]
        nonempty = list(dict.fromkeys(value for value in candidates if value))
        record[key] = nonempty[0] if nonempty else ""
        if len(nonempty) > 1:
            conflicts.append(key)
    for field in FIELD_ALIASES:
        record.setdefault(field, "")
    return record, conflicts


def _resolve_field_value(record: dict[str, str], field_or_alias: str) -> str:
    key = field_or_alias.strip()
    if not key:
        return ""
    if record.get(key, "").strip():
        return record[key].strip()
    canonical = _ALIAS_TO_CANONICAL.get(key)
    if canonical and record.get(canonical, "").strip():
        return record[canonical].strip()
    # Search aliases of canonical if key itself is canonical
    for alias, target in _ALIAS_TO_CANONICAL.items():
        if target == key and record.get(alias, "").strip():
            return record[alias].strip()
    return ""


def build_fill_values(
    placeholders: set[str],
    record: dict[str, str],
    placeholder_map: dict[str, str],
) -> dict[str, str]:
    """Map template placeholder text → fill string (keys = template-side names)."""
    values: dict[str, str] = {}
    for ph in placeholders:
        if ph in placeholder_map:
            source = placeholder_map[ph]
            values[ph] = _resolve_field_value(record, source)
            continue
        # Identity / alias / excel header
        resolved = _resolve_field_value(record, ph)
        if resolved:
            values[ph] = resolved
            continue
        # Placeholder is canonical english and record has it
        values[ph] = record.get(ph, "")
    # Also expose canonical keys for templates that mix both
    for field, value in record.items():
        values.setdefault(field, value)
    return values


def validate_record(
    record: dict[str, str],
    required_fields: list[str],
    conflicts: list[str] | None = None,
) -> tuple[list[str], list[str]]:
    missing = [field for field in required_fields if not _resolve_field_value(record, field)]
    return missing, sorted(conflicts or [])


def _replace_xml(xml_bytes: bytes, values: dict[str, str]) -> tuple[bytes, set[str]]:
    root = ElementTree.fromstring(xml_bytes)
    text_tag = f"{{{WORD_NS}}}t"
    paragraph_tag = f"{{{WORD_NS}}}p"
    for paragraph in root.iter(paragraph_tag):
        nodes = list(paragraph.iter(text_tag))
        if not nodes:
            continue
        original_nodes = [node.text or "" for node in nodes]
        combined = "".join(original_nodes)
        matches = list(PLACEHOLDER_RE.finditer(combined))
        if not matches:
            continue

        owners: list[int] = []
        for index, text in enumerate(original_nodes):
            owners.extend([index] * len(text))
        rendered_nodes = [""] * len(nodes)
        cursor = 0
        for match in matches:
            for position in range(cursor, match.start()):
                rendered_nodes[owners[position]] += combined[position]
            key = match.group(1).strip()
            # Key present (even empty) → replace; unknown key → leave token for unresolved report
            replacement = values[key] if key in values else match.group(0)
            owner = owners[match.start()] if match.start() < len(owners) else 0
            rendered_nodes[owner] += replacement
            cursor = match.end()
        for position in range(cursor, len(combined)):
            rendered_nodes[owners[position]] += combined[position]
        for node, text in zip(nodes, rendered_nodes):
            node.text = text
    rendered = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
    unresolved = set(PLACEHOLDER_RE.findall("".join(root.itertext())))
    # Normalise whitespace in unresolved names
    unresolved = {name.strip() for name in unresolved}
    return rendered, unresolved


def render_docx(template_path: Path, output_path: Path, values: dict[str, str]) -> list[str]:
    unresolved: set[str] = set()
    with zipfile.ZipFile(template_path, "r") as source, zipfile.ZipFile(output_path, "w") as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename.startswith("word/") and item.filename.endswith(".xml"):
                data, remaining = _replace_xml(data, values)
                unresolved.update(remaining)
            target.writestr(item, data)
    return sorted(unresolved)


def template_placeholders(template_path: Path) -> set[str]:
    placeholders: set[str] = set()
    with zipfile.ZipFile(template_path, "r") as source:
        for name in source.namelist():
            if name.startswith("word/") and name.endswith(".xml"):
                try:
                    root = ElementTree.fromstring(source.read(name))
                except ElementTree.ParseError:
                    continue
                placeholders.update(
                    match.strip() for match in PLACEHOLDER_RE.findall("".join(root.itertext()))
                )
    return placeholders


def _safe_filename(value: str, row_number: int) -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|\s]+", "_", value).strip("_")
    return cleaned or f"第{row_number}行"


def _excel_safe(value: object) -> object:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _write_report(rows: list[dict[str, object]], output_path: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as error:
        raise RuntimeError("缺少 openpyxl，无法读写 Excel；请使用桌面应用打包运行时") from error

    workbook = Workbook()
    log_sheet = workbook.active
    log_sheet.title = "生成日志"
    headers = [
        "来源行",
        "项目",
        "对公主体",
        "状态",
        "输出文件",
        "缺失字段",
        "冲突字段",
        "模板问题",
        "未替换占位符",
        "人工复核",
    ]
    log_sheet.append([_excel_safe(value) for value in headers])
    for row in rows:
        log_sheet.append([_excel_safe(row.get(header, "")) for header in headers])
    todo_sheet = workbook.create_sheet("待处理事项")
    todo_sheet.append([_excel_safe(value) for value in headers])
    for row in rows:
        if row.get("状态") != "已生成":
            todo_sheet.append([_excel_safe(row.get(header, "")) for header in headers])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _inspect(
    input_path: Path,
    template_path: Path,
    sheet_name: str | None,
    map_config: dict[str, object],
) -> dict[str, object]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("缺少 openpyxl，无法读写 Excel；请使用桌面应用打包运行时") from error

    placeholders = sorted(template_placeholders(template_path)) if template_path.exists() else []
    headers: list[str] = []
    if input_path.exists():
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        if sheet_name and sheet_name not in workbook.sheetnames:
            raise ValueError(f"输入工作表不存在：{sheet_name}")
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            first_row = next(rows)
            headers = [_display_value(value) for value in first_row]
        except StopIteration:
            headers = []
        workbook.close()

    suggested_placeholders: dict[str, str] = {}
    for ph in placeholders:
        if ph in map_config["placeholders"]:  # type: ignore[index]
            suggested_placeholders[ph] = map_config["placeholders"][ph]  # type: ignore[index]
        elif ph in _ALIAS_TO_CANONICAL:
            suggested_placeholders[ph] = _ALIAS_TO_CANONICAL[ph]
        elif ph in headers:
            suggested_placeholders[ph] = ph
        else:
            suggested_placeholders[ph] = ""

    return {
        "template_placeholders": placeholders,
        "excel_headers": headers,
        "suggested_placeholder_map": suggested_placeholders,
        "required_fields": map_config["required_fields"],
        "unmapped_placeholders": sorted(
            ph for ph, target in suggested_placeholders.items() if not target
        ),
    }


def _emit_deliverable(path: Path) -> str:
    """Print session product-card marker; prefer session-cwd-relative paths.

    Absolute paths under Application Support expert-sessions only resolve when
    the client passes the correct sessionRoot. Relative paths resolve reliably
    against the expert session directory and show on the turn product strip.
    """
    try:
        resolved = path.resolve()
        try:
            text = str(resolved.relative_to(Path.cwd().resolve())).replace("\\", "/")
        except ValueError:
            text = str(resolved)
    except OSError:
        text = str(path)
    print(f"ONMYAGENT_DELIVERABLE: {text}", flush=True)
    return text


def generate(args: argparse.Namespace) -> dict[str, object]:
    input_path = Path(args.input)
    template_path = Path(args.template)
    output_dir = Path(args.output_dir)
    report_path = Path(args.report)
    if template_path.suffix.lower() != ".docx" or not template_path.exists():
        raise ValueError("必须提供业务方批准的 DOCX 合同模板")
    if not input_path.exists():
        raise ValueError("对公返点信息 Excel 不存在")

    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("缺少 openpyxl，无法读写 Excel；请使用桌面应用打包运行时") from error

    raw_map = {}
    if args.map:
        raw_map = _load_json(args.map)
    elif args.column_map:
        raw_map = _load_json(args.column_map)
    map_config = _parse_map_config(raw_map)
    placeholder_map: dict[str, str] = map_config["placeholders"]  # type: ignore[assignment]
    required_fields: list[str] = map_config["required_fields"]  # type: ignore[assignment]
    custom_columns: dict[str, list[str]] = map_config["columns"]  # type: ignore[assignment]

    placeholders = template_placeholders(template_path)
    # Advisory only — never block batch solely because template uses CJK keys
    unmapped = sorted(
        ph
        for ph in placeholders
        if ph not in placeholder_map
        and ph not in _ALIAS_TO_CANONICAL
        and ph not in FIELD_ALIASES
    )
    template_notes = []
    if unmapped:
        template_notes.append("待确认映射：" + "、".join(f"{{{{{p}}}}}" for p in unmapped))

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    if args.sheet and args.sheet not in workbook.sheetnames:
        raise ValueError(f"输入工作表不存在：{args.sheet}")
    sheet = workbook[args.sheet] if args.sheet else workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        first_row = next(rows)
    except StopIteration as error:
        raise ValueError("输入工作表为空") from error
    headers = [_display_value(value) for value in first_row]
    mapping = _header_mapping(headers, custom_columns)
    output_dir.mkdir(parents=True, exist_ok=True)
    logs: list[dict[str, object]] = []
    generated = 0
    incomplete = 0
    deliverables: list[str] = []

    for row_number, values in enumerate(rows, start=2):
        if not any(_display_value(value) for value in values):
            continue
        record, conflicts = _normalise_record(list(values), mapping)
        missing, conflicts = validate_record(record, required_fields, conflicts)
        log: dict[str, object] = {
            "来源行": row_number,
            "项目": _resolve_field_value(record, "project_name") or record.get("项目", ""),
            "对公主体": _resolve_field_value(record, "counterparty_name")
            or record.get("主体名称", "")
            or record.get("公司名称", ""),
            "状态": "待补材料",
            "输出文件": "",
            "缺失字段": "、".join(missing),
            "冲突字段": "、".join(conflicts),
            "模板问题": "；".join(template_notes),
            "未替换占位符": "",
            "人工复核": "是",
        }
        if missing or conflicts:
            incomplete += 1
            logs.append(log)
            continue

        fill_values = build_fill_values(placeholders, record, placeholder_map)
        party = log["对公主体"] or f"第{row_number}行"
        filename = f"返点合同_{_safe_filename(str(party), row_number)}_{row_number}.docx"
        output_path = output_dir / filename
        unresolved = render_docx(template_path, output_path, fill_values)
        if unresolved:
            output_path.unlink(missing_ok=True)
            log["未替换占位符"] = "、".join(unresolved)
            incomplete += 1
        else:
            log["状态"] = "已生成"
            log["输出文件"] = filename
            generated += 1
            deliverables.append(_emit_deliverable(output_path))
        logs.append(log)

    workbook.close()
    _write_report(logs, report_path)
    if report_path.exists():
        deliverables.append(_emit_deliverable(report_path))
    return {
        "generated": generated,
        "incomplete": incomplete,
        "deliverables": deliverables,
    }


def self_test() -> dict[str, object]:
    complete = {field: "示例值" for field in FIELD_ALIASES}
    complete["rebate_amount"] = "1000"
    incomplete = dict(complete)
    incomplete["bank_account"] = ""
    complete_missing, complete_conflicts = validate_record(complete, list(DEFAULT_REQUIRED_FIELDS))
    incomplete_missing, _ = validate_record(incomplete, list(DEFAULT_REQUIRED_FIELDS))

    # English split across runs (existing contract)
    sample_en = (
        b'<?xml version="1.0" encoding="UTF-8"?>'
        b'<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        b"<w:body><w:p><w:r><w:t>{{counterparty_</w:t></w:r>"
        b"<w:r><w:t>name}}</w:t></w:r></w:p></w:body></w:document>"
    )
    _, unresolved_en = _replace_xml(sample_en, complete)

    # Chinese placeholders (officecli-compatible)
    sample_zh = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body><w:p><w:r><w:t>主体：{{甲方名称}} 金额：{{返点金额}}</w:t></w:r></w:p>"
        "</w:body></w:document>"
    ).encode("utf-8")
    zh_values = build_fill_values(
        {"甲方名称", "返点金额"},
        complete,
        {"甲方名称": "counterparty_name", "返点金额": "rebate_amount"},
    )
    rendered_zh, unresolved_zh = _replace_xml(sample_zh, zh_values)
    text_zh = ElementTree.fromstring(rendered_zh)
    body_text = "".join(text_zh.itertext())

    return {
        "complete_records": int(not complete_missing and not complete_conflicts),
        "incomplete_records": int(bool(incomplete_missing)),
        "unresolved_placeholders": sorted(unresolved_en),
        "chinese_placeholders_ok": int(
            "示例值" in body_text and "1000" in body_text and not unresolved_zh
        ),
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="根据批准的 DOCX 模板批量生成返点合同（officecli merge 的离线回退）"
    )
    parser.add_argument("--input", help="对公返点信息 Excel")
    parser.add_argument("--template", help="业务方批准的 DOCX 合同模板")
    parser.add_argument("--output-dir", help="合同输出目录")
    parser.add_argument("--report", help="生成报告 Excel")
    parser.add_argument("--sheet", help="输入工作表名称")
    parser.add_argument(
        "--map",
        help="统一映射 JSON（placeholders / columns / required_fields）；推荐",
    )
    parser.add_argument(
        "--column-map",
        help="兼容旧参数：列映射 JSON；可用 --map 替代",
    )
    parser.add_argument(
        "--inspect",
        action="store_true",
        help="只发现模板占位符与 Excel 表头，输出建议映射 JSON",
    )
    parser.add_argument("--self-test", action="store_true")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    if args.self_test:
        print(json.dumps(self_test(), ensure_ascii=False))
        return 0

    raw_map = {}
    if getattr(args, "map", None):
        try:
            raw_map = _load_json(args.map)
        except (OSError, ValueError, json.JSONDecodeError) as error:
            print(f"映射解析失败：{error}", file=sys.stderr)
            return 1
    elif args.column_map:
        try:
            raw_map = _load_json(args.column_map)
        except (OSError, ValueError, json.JSONDecodeError) as error:
            print(f"映射解析失败：{error}", file=sys.stderr)
            return 1
    map_config = _parse_map_config(raw_map)

    if args.inspect:
        if not args.input or not args.template:
            print("inspect 需要 --input 与 --template", file=sys.stderr)
            return 2
        try:
            result = _inspect(
                Path(args.input),
                Path(args.template),
                args.sheet,
                map_config,
            )
        except (OSError, ValueError, RuntimeError, zipfile.BadZipFile) as error:
            print(f"inspect 失败：{error}", file=sys.stderr)
            return 1
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0

    if not all((args.input, args.template, args.output_dir, args.report)):
        print("必须提供 --input、--template、--output-dir 和 --report", file=sys.stderr)
        return 2
    try:
        result = generate(args)
    except (OSError, ValueError, RuntimeError, zipfile.BadZipFile, json.JSONDecodeError) as error:
        print(f"生成失败：{error}", file=sys.stderr)
        return 1
    print(json.dumps(result, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
