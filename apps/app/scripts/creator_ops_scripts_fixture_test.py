#!/usr/bin/env python3
"""File-level fixture tests for the creator-ops deterministic scripts."""

from __future__ import annotations

import importlib.util
import json
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.dont_write_bytecode = True


REPO_ROOT = Path(__file__).resolve().parents[3]
EXPERTS_ROOT = REPO_ROOT / "apps/desktop/resources/marketplace/experts/plugins"
GENERATOR = EXPERTS_ROOT / "kol-content-ops-specialist/skills/rebate-contract-generator/scripts/generate_rebate_contracts.py"
WRITEBACK = EXPERTS_ROOT / "kol-content-ops-specialist/skills/rebate-contract-generator/scripts/writeback_completion_date.py"
CHECKER = EXPERTS_ROOT / "kol-content-ops-specialist/skills/rebate-contract-checker/scripts/check_rebate_contracts.py"
MARGIN = EXPERTS_ROOT / "kol-project-review-specialist/skills/kol-margin-effect-analysis/scripts/analyze_kol_performance.py"
MATRIX = EXPERTS_ROOT / "kol-project-review-specialist/skills/kol-content-performance-attribution/scripts/build_content_matrix.py"
XHS_DOCX = EXPERTS_ROOT / "kol-content-ops-specialist/skills/xhs-script-assistant/references/docx-template.py"


def run_script(script: Path, *args: object) -> None:
    result = subprocess.run(
        [sys.executable, str(script), *[str(arg) for arg in args]],
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode:
        raise AssertionError(result.stderr or result.stdout)


def create_minimal_docx(path: Path, *, chinese: bool = False) -> None:
    if chinese:
        body = """
    <w:p><w:r><w:rPr><w:b/></w:rPr><w:t>主体：{{甲方名称}}（保留格式）</w:t></w:r></w:p>
    <w:p><w:r><w:t>税号：{{统一社会信用代码}}</w:t></w:r></w:p>
    <w:p><w:r><w:t>账号：{{银行账号}}</w:t></w:r></w:p>
    <w:p><w:r><w:t>开户行：{{开户行}}</w:t></w:r></w:p>
    <w:p><w:r><w:t>发票内容：{{发票内容}}</w:t></w:r></w:p>
    <w:p><w:r><w:t>发票类型：{{发票类型}}</w:t></w:r></w:p>
    <w:p><w:r><w:t>金额：{{返点金额}}</w:t></w:r></w:p>
    <w:p><w:r><w:t>周期：{{合作周期}}</w:t></w:r></w:p>
"""
    else:
        body = """
    <w:p><w:r><w:rPr><w:b/></w:rPr><w:t>主体：{{counterparty_</w:t></w:r><w:r><w:rPr><w:i/></w:rPr><w:t>name}}（保留格式）</w:t></w:r></w:p>
    <w:p><w:r><w:t>税号：{{unified_social_credit_code}}</w:t></w:r></w:p>
    <w:p><w:r><w:t>账号：{{bank_account}}</w:t></w:r></w:p>
    <w:p><w:r><w:t>开户行：{{bank_name}}</w:t></w:r></w:p>
    <w:p><w:r><w:t>发票内容：{{invoice_content}}</w:t></w:r></w:p>
    <w:p><w:r><w:t>发票类型：{{invoice_type}}</w:t></w:r></w:p>
    <w:p><w:r><w:t>金额：{{rebate_amount}}</w:t></w:r></w:p>
    <w:p><w:r><w:t>周期：{{cooperation_period}}</w:t></w:r></w:p>
"""
    document = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:body>
{body}
    <w:sectPr/>
  </w:body>
</w:document>"""
    content_types = """<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>"""
    relationships = """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>"""
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("_rels/.rels", relationships)
        archive.writestr("word/document.xml", document)


def import_checker():
    spec = importlib.util.spec_from_file_location("rebate_checker", CHECKER)
    if spec is None or spec.loader is None:
        raise RuntimeError("cannot import rebate checker")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


class CreatorOpsScriptFixtures(unittest.TestCase):
    def test_completion_date_writeback_updates_session_copy_and_authorized_source(self):
        with tempfile.TemporaryDirectory(prefix="oma-completion-writeback-") as temp:
            root = Path(temp)
            source = root / "【ai】对公返点信息.xlsx"
            output = root / "session" / "【ai】对公返点信息_已填写完成日期.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["B1"] = "完成后填写日期"
            sheet.append(["序号", "完成日期", "博主名称"])
            for sequence in range(1, 7):
                sheet.append([sequence, None, f"达人{sequence}"])
            workbook.save(source)

            result = subprocess.run(
                [
                    sys.executable,
                    str(WRITEBACK),
                    "--input",
                    str(source),
                    "--output",
                    str(output),
                    "--date",
                    "2026-08-13",
                    "--source-writeback",
                ],
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, result.stderr or result.stdout)
            self.assertIn("ONMYAGENT_DELIVERABLE:", result.stdout)
            status = json.loads(result.stdout.splitlines()[0])
            self.assertTrue(status["source_updated"])
            self.assertEqual(status["rows_updated"], 6)
            self.assertEqual(status["cells"], ["B3", "B4", "B5", "B6", "B7", "B8"])
            for workbook_path in (source, output):
                updated = load_workbook(workbook_path, read_only=True, data_only=True)
                self.assertEqual(
                    [updated.active[f"B{row}"].value for row in range(3, 9)],
                    ["2026-08-13"] * 6,
                )
                updated.close()

    def test_completion_date_copy_does_not_mutate_source_without_authorization(self):
        with tempfile.TemporaryDirectory(prefix="oma-completion-copy-only-") as temp:
            root = Path(temp)
            source = root / "源台账.xlsx"
            output = root / "session" / "交付台账.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["序号", "完成日期"])
            sheet.append([1, None])
            workbook.save(source)

            result = subprocess.run(
                [
                    sys.executable,
                    str(WRITEBACK),
                    "--input",
                    str(source),
                    "--output",
                    str(output),
                    "--date",
                    "2026-08-13",
                ],
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, result.stderr or result.stdout)
            source_book = load_workbook(source, read_only=True, data_only=True)
            output_book = load_workbook(output, read_only=True, data_only=True)
            self.assertIsNone(source_book.active["B2"].value)
            self.assertEqual(output_book.active["B2"].value, "2026-08-13")
            source_book.close()
            output_book.close()

    def test_script_batch_renderer_writes_one_docx_per_creator(self):
        if importlib.util.find_spec("docx") is None:
            self.skipTest("python-docx is provided by the desktop packaged runtime")
        with tempfile.TemporaryDirectory(prefix="oma-script-fixture-") as temp:
            root = Path(temp)
            payload = {
                "project": "示例项目",
                "stage": "视频脚本",
                "reports": [
                    {
                        "talent": talent,
                        "product": "示例产品",
                        "script_rows": [["开头", "原稿", "留痕修改稿"]],
                        "issue_rows": [["开头", "绝对化表达", "存在绝对承诺", "P0", "改为客观体验", "更适合我"]],
                        "brief_check_rows": [["核心卖点", "满足"]],
                    }
                    for talent in ("达人A", "达人A")
                ],
            }
            source = root / "batch.json"
            source.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
            output_dir = root / "输出"
            output_dir.mkdir()
            run_script(XHS_DOCX, "--batch", source, output_dir)
            documents = sorted(output_dir.glob("*.docx"))
            self.assertEqual(len(documents), 2)
            self.assertEqual(len({document.name for document in documents}), 2)
            self.assertTrue(any(document.stem.endswith("_2") for document in documents))
            with zipfile.ZipFile(documents[0]) as archive:
                xml = archive.read("word/document.xml").decode("utf-8")
            self.assertIn("绝对化表达", xml)
            self.assertIn("P0", xml)

    def test_script_clean_renderer_excludes_audit_sections_and_rejects_markup(self):
        if importlib.util.find_spec("docx") is None:
            self.skipTest("python-docx is provided by the desktop packaged runtime")
        with tempfile.TemporaryDirectory(prefix="oma-script-clean-") as temp:
            root = Path(temp)
            clean_json = root / "clean.json"
            clean_output = root / "【视频脚本-示例产品-达人A】清洁终稿.docx"
            clean_json.write_text(
                json.dumps(
                    {
                        "title": "示例产品清洁口播终稿",
                        "meta": "达人：达人A | 平台：小红书",
                        "clean_paragraphs": [
                            "家人们，最近加班真的有点多。",
                            "这版只保留可以直接口播的正文。",
                        ],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            result = subprocess.run(
                [sys.executable, str(XHS_DOCX), "--clean", str(clean_json), str(clean_output)],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(result.returncode, 0, result.stderr or result.stdout)
            self.assertIn("ONMYAGENT_DELIVERABLE:", result.stdout)
            with zipfile.ZipFile(clean_output) as archive:
                xml = archive.read("word/document.xml").decode("utf-8")
            self.assertIn("家人们", xml)
            self.assertNotIn("审核核对", xml)
            self.assertNotIn("修改建议", xml)

            bad_json = root / "bad.json"
            bad_json.write_text(
                json.dumps({"clean_paragraphs": ["~~删除内容~~ <span>新增</span>"]}, ensure_ascii=False),
                encoding="utf-8",
            )
            bad = subprocess.run(
                [sys.executable, str(XHS_DOCX), "--clean", str(bad_json), str(root / "bad.docx")],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(bad.returncode, 2)
            self.assertIn("仍包含留痕", bad.stderr)

    def test_contract_generator_writes_only_complete_rows(self):
        with tempfile.TemporaryDirectory(prefix="oma-contract-fixture-") as temp:
            root = Path(temp)
            source = root / "返点信息.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append([
                "项目", "主体名称", "统一社会信用代码", "银行账号", "开户行",
                "发票内容", "发票类型", "返点金额", "合作周期",
            ])
            sheet.append(["=HYPERLINK(\"https://invalid\")", "示例公司", "91310000TEST", "62220000", "示例银行", "信息服务费", "普票", 1000, "2026-01-01 至 2026-01-31"])
            sheet["A2"].data_type = "s"
            sheet.append(["项目B", "缺字段公司", "91310000MISS", "", "示例银行", "信息服务费", "普票", 500, "2026-02"])
            workbook.save(source)
            template = root / "批准模板.docx"
            create_minimal_docx(template)
            output_dir = root / "合同"
            report = root / "生成报告.xlsx"
            result = subprocess.run(
                [
                    sys.executable,
                    str(GENERATOR),
                    "--input",
                    str(source),
                    "--template",
                    str(template),
                    "--output-dir",
                    str(output_dir),
                    "--report",
                    str(report),
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(result.returncode, 0, result.stderr or result.stdout)
            self.assertIn("ONMYAGENT_DELIVERABLE:", result.stdout)
            self.assertRegex(result.stdout, r"ONMYAGENT_DELIVERABLE:.*\.docx")
            self.assertNotIn("生成报告.xlsx", result.stdout)

            contracts = list(output_dir.glob("*.docx"))
            self.assertEqual(len(contracts), 1)
            with zipfile.ZipFile(contracts[0]) as archive:
                xml = archive.read("word/document.xml").decode("utf-8")
            self.assertIn("示例公司", xml)
            self.assertIn("（保留格式）", xml)
            self.assertRegex(xml, r"<[^>]*rPr><[^>]*i[^>]*/></[^>]*rPr><[^>]*t>（保留格式）")
            self.assertNotIn("{{", xml)
            report_book = load_workbook(report, read_only=False, data_only=False)
            self.assertEqual(report_book.sheetnames, ["生成日志", "待处理事项"])
            self.assertEqual(report_book["待处理事项"].max_row, 2)
            self.assertEqual(report_book["生成日志"]["B2"].data_type, "s")
            self.assertTrue(report_book["生成日志"]["B2"].value.startswith("'="))
            report_book.close()
            source_book = load_workbook(source, read_only=True, data_only=True)
            self.assertEqual(source_book.active.max_row, 3)
            self.assertIsNone(source_book.active["D3"].value)
            source_book.close()

    def test_contract_generator_fills_chinese_placeholders_with_map(self):
        with tempfile.TemporaryDirectory(prefix="oma-contract-zh-") as temp:
            root = Path(temp)
            source = root / "返点信息.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append([
                "项目", "主体名称", "统一社会信用代码", "银行账号", "开户行",
                "发票内容", "发票类型", "返点金额", "合作周期",
            ])
            sheet.append([
                "项目A", "中文模板公司", "91310000ZH", "62221111", "中文银行",
                "信息服务费", "普票", 2000, "2026-03",
            ])
            workbook.save(source)
            template = root / "中文模板.docx"
            create_minimal_docx(template, chinese=True)
            mapping = root / "mapping.json"
            mapping.write_text(
                json.dumps(
                    {
                        "placeholders": {
                            "甲方名称": "counterparty_name",
                            "统一社会信用代码": "unified_social_credit_code",
                            "银行账号": "bank_account",
                            "开户行": "bank_name",
                            "发票内容": "invoice_content",
                            "发票类型": "invoice_type",
                            "返点金额": "rebate_amount",
                            "合作周期": "cooperation_period",
                        }
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            output_dir = root / "合同"
            report = root / "生成报告.xlsx"
            run_script(
                GENERATOR,
                "--input",
                source,
                "--template",
                template,
                "--output-dir",
                output_dir,
                "--report",
                report,
                "--map",
                mapping,
            )
            contracts = list(output_dir.glob("*.docx"))
            self.assertEqual(len(contracts), 1)
            with zipfile.ZipFile(contracts[0]) as archive:
                xml = archive.read("word/document.xml").decode("utf-8")
            self.assertIn("中文模板公司", xml)
            self.assertIn("2000", xml)
            self.assertNotIn("{{", xml)

    def test_margin_workbook_keeps_formulas_and_data_gaps_traceable(self):
        with tempfile.TemporaryDirectory(prefix="oma-margin-fixture-") as temp:
            root = Path(temp)
            source = root / "效果.xlsx"
            output = root / "毛利效果.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["记录ID", "达人", "返点档", "毛利率", "达人量级", "笔记类型", "K金额", "广告金额", "自然阅读量", "自然曝光量", "点赞量", "评论量", "收藏量", "搜索进店UV", "K 金额"])
            sheet.append(["note-1", "=CMD()", "10%", "30%", "腰部", "测评", 1000, 200, 100, 1000, 10, 5, 5, 10, 1000])
            sheet["B2"].data_type = "s"
            sheet.append(["note-2", "达人B", "15%", "25%", "尾部", "教程", 500, 0, 0, 0, 2, 1, 1, 0, 500])
            sheet.append(["note-dup", "达人C", "10%", "20%", "尾部", "测评", 300, 0, 30, 300, 1, 1, 1, 1, 300])
            sheet.append(["note-dup", "达人C", "10%", "20%", "尾部", "测评", 300, 0, 30, 300, 1, 1, 1, 1, 300])
            sheet.append(["note-conflict", "达人D", "10%", "20%", "尾部", "教程", 1000, 0, 50, 500, 1, 1, 1, 1, 999])
            workbook.save(source)
            run_script(MARGIN, "--input", source, "--output", output)
            result = load_workbook(output, read_only=False, data_only=False)
            self.assertEqual(result.sheetnames, ["指标明细", "口径表", "分组分析", "可复投建议", "风险提示", "清洗日志"])
            detail = list(result["指标明细"].iter_rows(values_only=True))
            headers = list(detail[0])
            first = dict(zip(headers, detail[1]))
            second = dict(zip(headers, detail[2]))
            self.assertEqual(first["总金额"], 1200)
            self.assertEqual(first["自然 CTR"], 0.1)
            self.assertEqual(first["纯 K CPC"], 10)
            self.assertEqual(first["行级标签"], "数据不足")
            self.assertEqual(result["指标明细"]["C2"].data_type, "s")
            self.assertTrue(result["指标明细"]["C2"].value.startswith("'="))
            recommendation_text = "\n".join(
                str(cell)
                for row in result["可复投建议"].iter_rows(values_only=True)
                for cell in row
            )
            self.assertIn("补齐数据或业务阈值后再判断", recommendation_text)
            self.assertEqual(second["自然 CTR"], "数据不足")
            self.assertIn("源行", headers)
            risk_text = "\n".join(str(cell) for row in result["风险提示"].iter_rows(values_only=True) for cell in row)
            self.assertIn("重复记录", risk_text)
            self.assertIn("字段冲突", risk_text)
            result.close()

    def test_margin_two_source_mode_exact_matches_links_and_routes_exceptions(self):
        with tempfile.TemporaryDirectory(prefix="oma-margin-two-source-") as temp:
            root = Path(temp)
            plan = root / "项目计划单.xlsx"
            effect = root / "蒲公英效果.xlsx"
            output = root / "毛利效果分析.xlsx"

            plan_book = Workbook()
            plan_sheet = plan_book.active
            plan_sheet.title = "返点毛利计划单"
            plan_sheet.append(["达人昵称", "笔记链接", "粉丝量级", "笔记类型", "K金额", "返点档", "毛利率"])
            plan_sheet.append(["计划昵称A", "https://example.test/a", "10-20万", "图文", 1000, "<20%", 0.2])
            plan_sheet.append(["计划昵称B", "https://example.test/b", "20-50万", "视频", 2000, "20%-25%", 0.25])
            plan_sheet.append(["重复链接", "https://example.test/dup", "5-10万", "图文", 500, "<20%", 0.1])
            plan_sheet.append(["只在计划单", "https://example.test/plan-only", "5-10万", "图文", 500, "<20%", 0.1])
            method = plan_book.create_sheet("分析口径与阈值")
            method.append(["类型", "指标/规则", "公式或口径", "演示阈值", "说明"])
            method.append(["纯K前端", "自然CTR", "自然阅读量 / 自然曝光量", ">=8.0%", "零分母标数据不足"])
            method.append(["纯K前端", "纯K CPC", "K金额 / 自然阅读量", "<=¥12", ""])
            method.append(["总投入", "搜索进店成本", "总金额 / 搜索进店UV", "<=¥30", ""])
            plan_book.save(plan)

            effect_book = Workbook()
            effect_sheet = effect_book.active
            effect_sheet.title = "蒲公英投放效果"
            effect_sheet.append(["效果记录ID", "博主昵称", "发布链接", "自然曝光量", "自然阅读量", "广告金额", "点赞量", "评论量", "收藏量", "搜索进店UV", "推广曝光量", "推广阅读量"])
            effect_sheet.append(["EFF-A", "效果昵称不同", "https://example.test/a", 1000, 100, 200, 10, 5, 5, 50, 500, 60])
            effect_sheet.append(["EFF-B", "计划昵称B", "https://example.test/b", 0, 80, 100, 8, 4, 3, None, 300, 40])
            effect_sheet.append(["EFF-D1", "重复链接", "https://example.test/dup", 500, 50, 0, 2, 1, 1, 10, 100, 10])
            effect_sheet.append(["EFF-D2", "重复链接", "https://example.test/dup", 600, 60, 0, 3, 1, 1, 12, 120, 12])
            effect_sheet.append(["EFF-ONLY", "只在效果表", "https://example.test/effect-only", 300, 30, 0, 1, 1, 1, 5, 50, 5])
            effect_book.save(effect)

            result = subprocess.run(
                [
                    sys.executable,
                    str(MARGIN),
                    "--plan-input",
                    str(plan),
                    "--effect-input",
                    str(effect),
                    "--output",
                    str(output),
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(result.returncode, 0, result.stderr or result.stdout)
            self.assertIn("ONMYAGENT_DELIVERABLE:", result.stdout)
            self.assertIn("毛利效果分析.xlsx", result.stdout)

            workbook = load_workbook(output, read_only=False, data_only=False)
            self.assertEqual(workbook.sheetnames, ["清洗底表", "字段映射", "分组分析", "分析结论", "口径表", "人工确认", "清洗日志"])
            details = list(workbook["清洗底表"].iter_rows(values_only=True))
            headers = list(details[0])
            rows = [dict(zip(headers, row)) for row in details[1:]]
            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[0]["计划单达人昵称"], "计划昵称A")
            self.assertEqual(rows[0]["效果表博主昵称"], "效果昵称不同")
            self.assertEqual(rows[0]["行级标签"], "优先保留")
            self.assertEqual(rows[1]["行级标签"], "数据不足")
            manual_text = "\n".join(
                str(cell)
                for row in workbook["人工确认"].iter_rows(values_only=True)
                for cell in row
            )
            self.assertIn("不是一对一", manual_text)
            self.assertIn("计划单中未匹配", manual_text)
            self.assertIn("效果表中未匹配", manual_text)
            self.assertIn("自然 CTR", manual_text)
            self.assertIn("搜索进店成本", manual_text)
            mapping_text = "\n".join(
                str(cell)
                for row in workbook["字段映射"].iter_rows(values_only=True)
                for cell in row
            )
            self.assertIn("笔记链接", mapping_text)
            self.assertIn("仅供人工核对，不参与匹配", mapping_text)
            method_text = "\n".join(
                str(cell)
                for row in workbook["口径表"].iter_rows(values_only=True)
                for cell in row
            )
            self.assertIn("自然阅读量 / 自然曝光量", method_text)
            workbook.close()

    def test_content_matrices_use_correlation_language(self):
        with tempfile.TemporaryDirectory(prefix="oma-matrix-fixture-") as temp:
            root = Path(temp)
            source = root / "投放.xlsx"
            output = root / "内容矩阵.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["记录ID", "周次", "内容类型", "内容切角", "目标人群", "搜索词", "投入", "阅读量", "互动量", "笔记类型", "优化动作"])
            sheet.append(["a", "本周", "测评", "成分", "新客", "成分", 100, 1000, 100, "测评", "提高新客测试占比"])
            sheet.append(["b", "上周", "教程", "用法", "熟客", "用法", 120, 800, 80, "教程", "维持基线"])
            sheet.append(["dup", "本周", "测评", "成分", "新客", "成分", 20, 100, 10, "测评", ""])
            sheet.append(["dup", "本周", "测评", "成分", "新客", "成分", 20, 100, 10, "测评", ""])
            sheet.append(["conflict", "本周", "测评", "成分", "新客", "成分", 20, 100, 10, "教程", ""])
            sheet.append(["missing", "本周", "教程", "场景", "新客", "使用", None, 300, 30, "教程", ""])
            sheet.append(["hostile", "本周", "=CMD()", "场景", "新客", "使用", 10, 300, 30, "=CMD()", ""])
            sheet["C8"].data_type = "s"
            sheet["J8"].data_type = "s"
            workbook.save(source)
            run_script(MATRIX, "--input", source, "--output", output)
            result = load_workbook(output, read_only=False, data_only=False)
            self.assertEqual(result.sheetnames, ["周环比", "内容人群矩阵", "搜索内容矩阵", "内容切角", "风险提示"])
            content_rows = list(result["内容人群矩阵"].iter_rows(values_only=True))
            search_rows = list(result["搜索内容矩阵"].iter_rows(values_only=True))
            self.assertEqual(len(content_rows), 4)
            self.assertEqual(len(search_rows), 4)
            self.assertTrue(all(row[-2] == "相关性观察" for row in content_rows[1:]))
            weekly_text = "\n".join(
                str(cell)
                for row in result["周环比"].iter_rows(values_only=True)
                for cell in row
            )
            self.assertIn("提高新客测试占比", weekly_text)
            self.assertIn("源行", weekly_text)
            risk_text = "\n".join(str(cell) for row in result["风险提示"].iter_rows(values_only=True) for cell in row)
            self.assertIn("相关性不等于因果", risk_text)
            self.assertIn("重复记录", risk_text)
            self.assertIn("字段冲突", risk_text)
            self.assertIn("数值缺失或无效", risk_text)
            hostile_cells = [cell for row in result["内容人群矩阵"].iter_rows() for cell in row if isinstance(cell.value, str) and "CMD" in cell.value]
            self.assertTrue(hostile_cells)
            self.assertTrue(all(cell.data_type != "f" and cell.value.startswith("'=") for cell in hostile_cells))
            result.close()

    def test_checker_routes_duplicate_and_fuzzy_matches_to_manual_sheet(self):
        module = import_checker()
        checker = module.RebateContractChecker(amount_tolerance=1)
        invoice = {
            "来源文件": "invoice.xlsx", "项目": "=CMD()", "服务项目名称": "信息服务",
            "发票抬头": "示例公司", "税号": "TAX", "开票金额": 1000,
            "发票类型": "普通发票", "开票内容": "信息服务费",
        }
        contract = {
            "来源文件": "contract.pdf", "甲方名称": "示例公司", "税号": "TAX",
            "渠道服务费金额": 1001, "发票类型": "普通发票", "开票内容": "信息服务费",
            "博主列表": [{"博主名称": "达人A", "合作金额": 8000, "渠道服务费": 1000}],
        }
        unique = checker.match_invoice_to_contract([invoice], {"contract.pdf": contract})
        unique = checker.enrich_with_plan_data(unique, {"plan.xlsx": [{"博主名称": "达人A", "计划单返点金额": 1000, "付款方式": "对公", "日期": "2026-01", "来源文件": "plan.xlsx"}]})
        unique[0]["核对结果"] = checker.determine_final_result(unique[0])
        self.assertTrue(unique[0]["核对结果"].startswith("✅"))

        duplicates = checker.match_invoice_to_contract([invoice], {
            "a.pdf": dict(contract, 来源文件="a.pdf"),
            "b.pdf": dict(contract, 来源文件="b.pdf"),
        })
        for row in duplicates:
            row["核对结果"] = checker.determine_final_result(row)
        self.assertTrue(all("需人工确认" in row["核对结果"] for row in duplicates))

        missing_contract_detail = checker.match_invoice_to_contract(
            [invoice], {"contract.pdf": dict(contract, 博主列表=[])}
        )[0]
        missing_contract_detail["核对结果"] = checker.determine_final_result(missing_contract_detail)
        self.assertIn("三方核对不完整", missing_contract_detail["核对结果"])
        self.assertFalse(missing_contract_detail["核对结果"].startswith("✅"))

        missing_plan = checker.match_invoice_to_contract([invoice], {"contract.pdf": contract})
        checker.enrich_with_plan_data(missing_plan, {})
        missing_plan[0]["核对结果"] = checker.determine_final_result(missing_plan[0])
        self.assertIn("计划单中未找到", missing_plan[0]["核对结果"])
        self.assertFalse(missing_plan[0]["核对结果"].startswith("✅"))

        with tempfile.TemporaryDirectory(prefix="oma-checker-fixture-") as temp:
            report = Path(temp) / "核对报告.xlsx"
            checker.save_results(unique + duplicates + [missing_contract_detail, *missing_plan], report)
            workbook = load_workbook(report, read_only=False, data_only=False)
            self.assertEqual(workbook.sheetnames, ["核对清单", "核对摘要", "待处理事项"])
            self.assertEqual(workbook["待处理事项"].max_row, 5)
            self.assertEqual(workbook["核对清单"]["B2"].data_type, "s")
            self.assertTrue(workbook["核对清单"]["B2"].value.startswith("'="))
            summary = dict(workbook["核对摘要"].iter_rows(min_row=2, values_only=True))
            self.assertEqual(summary["金额容差（元）"], 1)
            workbook.close()


if __name__ == "__main__":
    unittest.main()
